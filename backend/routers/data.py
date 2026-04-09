"""Data router — economics, daily summary, fuel prices, sales, blackout, generators, site detail, upload."""
import os
import tempfile
from typing import Optional
from fastapi import APIRouter, Depends, UploadFile, File, Query, HTTPException
import pandas as pd
from utils.database import get_db, log_activity
from models.energy_cost import get_store_economics, get_generator_detail, get_trends
from backend.routers.auth import get_current_user

router = APIRouter()

# ─── Expected columns per file type — used for validation ─────────
EXPECTED_COLUMNS = {
    "daily_sales": {
        "required": ["SALES_DATE", "GOLD_CODE", "CostCenter", "SegmentName", "SALES_AMT", "MARGIN"],
    },
    "hourly_sales": {
        "required": ["DocumentDate", "Sales_HR", "GOLD_CODE", "CostCenter", "SegmentName", "TotalAmount"],
    },
    "store_master": {
        "required": ["GOLD_Code", "CostCenter", "SegmentName"],
    },
    "blackout": {
        "required": [],  # Dynamic date columns, validated by parser
    },
    "fuel_price": {
        "required": [],  # Multi-sheet, validated by parser
    },
    "diesel_expense": {
        "required": [],  # Flexible format, validated by parser
    },
}

# Known sheet names that the system processes
KNOWN_SHEETS = {
    "daily sales", "daily_sales", "hourly sales", "hourly_sales",
    "store master", "store_master", "storemaster",
    "all cmhl-sites", "cp", "cmhl", "cfc", "pg",
    "mapping", "sheet1", "data",
}


def _check_cols(found_cols, expected_cols, sheet_label, errors):
    """Check if expected columns exist in found columns (case-insensitive, ignore spaces/underscores)."""
    found_lower = [c.lower().replace(" ", "").replace("_", "") for c in found_cols]
    missing = []
    for exp in expected_cols:
        exp_clean = exp.lower().replace(" ", "").replace("_", "")
        if not any(exp_clean in f for f in found_lower):
            missing.append(exp)
    if missing:
        errors.append(
            f"{sheet_label}: MISSING COLUMNS — expected [{', '.join(expected_cols)}] "
            f"but could not find [{', '.join(missing)}]. "
            f"Found columns: [{', '.join(found_cols[:10])}]"
        )


def _validate_columns(filepath, file_type, sheets):
    """Validate Excel column headers match expected format. Returns (errors, warnings)."""
    import openpyxl

    warnings = []
    errors = []

    wb = openpyxl.load_workbook(filepath, read_only=True)

    # Check for unknown sheets
    for sn in sheets:
        if sn.lower().strip() not in KNOWN_SHEETS:
            warnings.append(f"UNKNOWN_SHEET: '{sn}' — will be ignored")

    # Validate columns for known sheet types
    if file_type in ("combo_sales", "daily_sales", "sales_reference"):
        for sn in sheets:
            sn_lower = sn.lower().strip()

            if "daily" in sn_lower and "sales" in sn_lower:
                ws = wb[sn]
                header_row = next(ws.iter_rows(max_row=1, values_only=True), None)
                if header_row:
                    found_cols = [str(c).strip() for c in header_row if c is not None]
                    expected = EXPECTED_COLUMNS["daily_sales"]["required"]
                    _check_cols(found_cols, expected, f"Sheet '{sn}'", errors)

            elif "hourly" in sn_lower and "sales" in sn_lower:
                ws = wb[sn]
                header_row = next(ws.iter_rows(max_row=1, values_only=True), None)
                if header_row:
                    found_cols = [str(c).strip() for c in header_row if c is not None]
                    expected = EXPECTED_COLUMNS["hourly_sales"]["required"]
                    _check_cols(found_cols, expected, f"Sheet '{sn}'", errors)

            elif sn_lower in ("store master", "store_master", "storemaster") or ("all" in sn_lower and "sites" in sn_lower):
                ws = wb[sn]
                header_row = next(ws.iter_rows(max_row=1, values_only=True), None)
                if header_row:
                    found_cols = [str(c).strip() for c in header_row if c is not None]
                    expected = EXPECTED_COLUMNS["store_master"]["required"]
                    _check_cols(found_cols, expected, f"Sheet '{sn}'", errors)

    elif file_type == "store_master":
        for sn in sheets:
            sn_lower = sn.lower().strip()
            if sn_lower in ("store master", "store_master", "storemaster", "sheet1", "data") or ("all" in sn_lower and "sites" in sn_lower):
                ws = wb[sn]
                header_row = next(ws.iter_rows(max_row=1, values_only=True), None)
                if header_row:
                    found_cols = [str(c).strip() for c in header_row if c is not None]
                    expected = EXPECTED_COLUMNS["store_master"]["required"]
                    _check_cols(found_cols, expected, f"Sheet '{sn}'", errors)

    wb.close()
    return errors, warnings


def _parse_coord(val):
    """Parse lat/lng — decimal, DMS (16°51'56.4"), or None."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    import re
    s = str(val).strip()
    if s in ("", "NULL", "null", "None"):
        return None
    try:
        return float(s)
    except ValueError:
        pass
    m = re.match(r"(\d+)[°](\d+)['\u2019](\d+\.?\d*)[\"″\u201d]?\s*([NSEW])?", s)
    if m:
        d, mi, sec = float(m.group(1)), float(m.group(2)), float(m.group(3))
        result = d + mi / 60 + sec / 3600
        if m.group(4) in ("S", "W"):
            result = -result
        return round(result, 8)
    return None


# ─── Helper: date filter SQL ─────────────────────────────
def _dsql(col: str, date_from: str = None, date_to: str = None):
    sql = ""
    params = []
    if date_from:
        sql += f" AND {col} >= ?"
        params.append(date_from)
    if date_to:
        sql += f" AND {col} <= ?"
        params.append(date_to)
    return sql, params


def _sector_price_on_date(conn, sector_id, date):
    """Latest purchase price for a sector on or before a given date."""
    row = conn.execute("""
        SELECT price_per_liter FROM fuel_purchases
        WHERE sector_id = ? AND date <= ? AND price_per_liter IS NOT NULL
        ORDER BY date DESC LIMIT 1
    """, (sector_id, date)).fetchone()
    return row[0] if row else 0


def _sector_prices_latest(conn):
    """Latest fuel price per sector (most recent purchase date)."""
    rows = conn.execute("""
        SELECT sector_id, price_per_liter, date FROM fuel_purchases
        WHERE price_per_liter IS NOT NULL
        GROUP BY sector_id HAVING date = MAX(date)
    """).fetchall()
    return {r["sector_id"]: r["price_per_liter"] for r in rows}


# ─── 1.6 Upload ──────────────────────────────────────────
@router.post("/upload/validate")
async def validate_file(file: UploadFile = File(...), user: dict = Depends(get_current_user)):
    """Validate file type without importing — returns detected type and sheets."""
    if user["role"] not in ("super_admin", "admin"):
        raise HTTPException(403, "Upload requires admin role")
    suffix = os.path.splitext(file.filename)[1]
    with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as tmp:
        content = await file.read()
        tmp.write(content)
        tmp_path = tmp.name
    try:
        import openpyxl
        from itertools import islice

        wb = openpyxl.load_workbook(tmp_path, read_only=True, data_only=True)
        sheets = wb.sheetnames
        sheets_lower = [s.lower().strip() for s in sheets]

        # Get row counts from XML metadata — no cell parsing needed
        sheet_rows = {}
        for sn in sheets:
            ws = wb[sn]
            sheet_rows[sn] = max((ws.max_row or 1) - 1, 0)  # subtract header

        # Detect type by sheet names
        detected_type = None

        if any(s in sheets_lower for s in ["daily sales", "daily_sales", "hourly sales", "hourly_sales"]):
            # Use metadata row count to classify (no pandas read needed)
            for sn in sheets:
                if "daily" in sn.lower() and "sales" in sn.lower():
                    detected_type = "sales_reference" if sheet_rows.get(sn, 0) > 10000 else "combo_sales"
                    break
            if not detected_type:
                detected_type = "combo_sales"
        elif len(sheets) >= 3 and sum(s in sheets_lower for s in ["cmhl", "cp", "cfc", "pg"]) >= 3:
            detected_type = "fuel_price"
        elif "cp" in sheets_lower:
            detected_type = "blackout_cp"
        elif "cmhl" in sheets_lower:
            detected_type = "blackout_cmhl"
        elif "cfc" in sheets_lower:
            detected_type = "blackout_cfc"
        elif "pg" in sheets_lower:
            detected_type = "blackout_pg"
        else:
            # Read first 2 rows from openpyxl (not pandas) — near instant
            try:
                ws = wb[sheets[0]]
                first_rows = list(islice(ws.iter_rows(max_row=2, values_only=True), 2))
                cols_text = ' '.join(str(c).lower() for c in first_rows[0] if c) if first_rows else ''
                vals_text = ' '.join(str(v).lower() for v in first_rows[1] if v) if len(first_rows) > 1 else ''
                all_text = cols_text + ' ' + vals_text

                if "store contribution" in all_text or ("store" in cols_text and "center" in cols_text and "cost center" in cols_text):
                    detected_type = "store_allocation"
                elif "daily average diesel" in all_text or "daily normal diesel" in all_text or "diesel expense" in all_text:
                    detected_type = "diesel_expense_ly"
                elif "daily sales" in all_text or "sales_amt" in all_text:
                    detected_type = "daily_sales"
                elif "store exp" in all_text or "expense percentage" in all_text:
                    detected_type = "diesel_expense_ly"
                elif "gold_code" in all_text or "goldcode" in all_text or "store master" in all_text:
                    detected_type = "store_master"
            except Exception:
                pass

        wb.close()

        if not detected_type:
            raise HTTPException(400, f"UNKNOWN_FILE_TYPE — sheets: {', '.join(sheets)}")

        # Row estimate from metadata (already collected above — no extra reads)
        row_estimate = 0
        for sn in sheets:
            sn_lower = sn.lower().strip()
            if sn_lower in ("cmhl", "cp", "cfc", "pg") or "daily" in sn_lower or "sales" in sn_lower or "hourly" in sn_lower or sn_lower in ("sheet1", "data"):
                row_estimate += sheet_rows.get(sn, 0)
        if row_estimate == 0:
            row_estimate = sheet_rows.get(sheets[0], 0)

        # Validate column headers before returning
        col_errors, col_warnings = _validate_columns(tmp_path, detected_type, sheets)

        return {
            "type": detected_type,
            "sheets": sheets,
            "filename": file.filename,
            "rows": row_estimate,
            "validation": col_errors + col_warnings,
            "valid": len(col_errors) == 0,
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(400, f"VALIDATION_ERROR: {e}")
    finally:
        os.unlink(tmp_path)


@router.post("/upload")
async def upload_file(file: UploadFile = File(...), file_type: Optional[str] = Query(None), user: dict = Depends(get_current_user)):
    """Upload and import Excel file. Optional file_type skips re-detection (from validate endpoint)."""
    if user["role"] not in ("super_admin", "admin"):
        raise HTTPException(403, "Upload requires admin role")

    suffix = os.path.splitext(file.filename)[1]
    with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as tmp:
        content = await file.read()
        tmp.write(content)
        tmp_path = tmp.name

    try:
        import openpyxl
        wb = openpyxl.load_workbook(tmp_path, read_only=True)
        sheets = wb.sheetnames
        wb.close()

        sheets_lower = [s.lower().strip() for s in sheets]
        detected_type = None
        record_count = 0

        # ─── Pre-detect file type for column validation ────────────
        # Use hint from validate endpoint if available, otherwise auto-detect
        _pre_type = None
        if file_type:
            # Map frontend file_type hints to internal types
            _type_map = {
                "blackout_cp": "blackout", "blackout_cmhl": "blackout",
                "blackout_cfc": "blackout", "blackout_pg": "blackout",
                "combo_sales": "combo_sales", "daily_sales": "daily_sales",
                "sales_reference": "combo_sales", "fuel_price": "fuel_price",
                "store_master": "store_master", "diesel_expense_ly": "diesel_expense",
                "store_allocation": "store_allocation",
            }
            _pre_type = _type_map.get(file_type, file_type)
        if not _pre_type:
            if any(s in sheets_lower for s in ["daily sales", "daily_sales", "hourly sales", "hourly_sales"]):
                _pre_type = "combo_sales"
            elif len(sheets) >= 3 and sum(s in sheets_lower for s in ["cmhl", "cp", "cfc", "pg"]) >= 3:
                _pre_type = "fuel_price"
            elif any(s in sheets_lower for s in ["cp", "cmhl", "cfc", "pg"]):
                _pre_type = "blackout"
            else:
                # Peek at first sheet headers
                try:
                    import openpyxl as _oxl_peek
                    from itertools import islice as _islice
                    _wb_peek = _oxl_peek.load_workbook(tmp_path, read_only=True, data_only=True)
                    _ws_peek = _wb_peek[sheets[0]]
                    _first = list(_islice(_ws_peek.iter_rows(max_row=2, values_only=True), 2))
                    _peek_text = ' '.join(str(c).lower() for c in _first[0] if c) if _first else ''
                    if "store contribution" in _peek_text or ("store" in _peek_text and "center" in _peek_text and "cost center" in _peek_text):
                        _pre_type = "store_allocation"
                    elif "gold_code" in _peek_text or "goldcode" in _peek_text or "store master" in _peek_text:
                        _pre_type = "store_master"
                    elif "sales_amt" in _peek_text or "daily sales" in _peek_text:
                        _pre_type = "daily_sales"
                    elif "diesel" in _peek_text or "store exp" in _peek_text:
                        _pre_type = "diesel_expense"
                    _wb_peek.close()
                except Exception:
                    pass

        # ─── Validate columns — block import if required columns missing ──
        col_errors, col_warnings = _validate_columns(tmp_path, _pre_type or "unknown", sheets)
        if col_errors:
            try:
                with get_db() as _log_conn:
                    log_activity(_log_conn, user["id"], user["username"], "UPLOAD_FAILED", "DATA", f"Upload rejected: {file.filename} — VALIDATION_FAILED", {"filename": file.filename, "errors": col_errors})
            except Exception:
                pass
            return {
                "filename": file.filename,
                "sheets": sheets,
                "records": 0,
                "type": _pre_type or "unknown",
                "error": "VALIDATION_FAILED",
                "validation": col_errors + col_warnings,
            }

        from utils.database import (
            get_db as _get_db, upsert_site, upsert_generator, upsert_daily_operation,
            insert_fuel_purchase, refresh_site_summary, log_upload,
            upsert_daily_sale, upsert_hourly_sale, upsert_diesel_expense_ly,
            enrich_sites_from_store_master,
        )

        # ─── FULL CLEAR before import — each upload is a complete data refresh ──
        with _get_db() as conn:
            if _pre_type == "combo_sales" or _pre_type == "daily_sales" or _pre_type == "store_master":
                conn.execute("DELETE FROM daily_sales")
                conn.execute("DELETE FROM hourly_sales")
                conn.execute("DELETE FROM store_master")
            elif _pre_type == "fuel_price":
                conn.execute("DELETE FROM fuel_purchases")
            elif _pre_type == "blackout":
                # Clear only the SPECIFIC sector being uploaded
                _sector = next((s for s in ["CP", "CMHL", "CFC", "PG"] if s.lower() in sheets_lower), None)
                if _sector:
                    conn.execute("DELETE FROM daily_operations WHERE site_id IN (SELECT site_id FROM sites WHERE sector_id=?)", (_sector,))
                    conn.execute("DELETE FROM daily_site_summary WHERE site_id IN (SELECT site_id FROM sites WHERE sector_id=?)", (_sector,))
                    conn.execute("DELETE FROM generators WHERE site_id IN (SELECT site_id FROM sites WHERE sector_id=?)", (_sector,))
                    conn.execute("DELETE FROM sites WHERE sector_id=?", (_sector,))
                    conn.execute("DELETE FROM excel_validation_cache WHERE sector_id=?", (_sector,))
            elif _pre_type == "diesel_expense":
                conn.execute("DELETE FROM diesel_expense_ly")
            elif _pre_type == "store_allocation":
                conn.execute("DELETE FROM store_center_allocation")
            # Clear derived/cached data only for data types that affect them
            if _pre_type in ("blackout", "combo_sales", "daily_sales", "fuel_price"):
                conn.execute("DELETE FROM alerts")
            conn.execute("DELETE FROM ai_insights_cache")
            try:
                log_activity(conn, user["id"], user["username"], "CLEAR", "DATA", f"Cleared data before {_pre_type} upload")
            except Exception:
                pass

        # ─── FUEL PRICE (must check BEFORE blackout — both have sector sheets) ──
        if len(sheets) >= 3 and sum(s in sheets_lower for s in ["cmhl", "cp", "cfc", "pg"]) >= 3 and not any(s in sheets_lower for s in ["daily sales", "daily_sales"]):
            from parsers.fuel_price_parser import parse_fuel_price_file
            parsed = parse_fuel_price_file(tmp_path)
            purchases = parsed.get("purchases", []) if isinstance(parsed, dict) else []
            detected_type = "fuel_price"

            with _get_db() as conn:
                log_upload(conn, file.filename, "fuel_price", None,
                           len(purchases), len(purchases), 0, None, None, None)
                for fp in purchases:
                    insert_fuel_purchase(conn, fp["sector_id"], fp["date"],
                                         fp.get("region"), fp.get("supplier"),
                                         fp.get("fuel_type"), fp.get("quantity_liters"),
                                         fp.get("price_per_liter"), source="api")
            record_count = len(purchases)

        # ─── BLACKOUT FILES ────────────────────────────────────
        elif any(s in sheets_lower for s in ["cp", "cmhl", "cfc", "pg"]) and not any(s in sheets_lower for s in ["daily sales", "daily_sales"]):
            from parsers.blackout_parser import parse_blackout_file
            sector = next((s for s in ["CP", "CMHL", "CFC", "PG"] if s.lower() in sheets_lower), None)
            if sector:
                parsed = parse_blackout_file(tmp_path, sector)
                gens = parsed.get("generators", [])
                daily = parsed.get("daily_data", [])
                dates = parsed.get("dates_found", [])
                errors = parsed.get("errors", [])
                detected_type = f"blackout_{sector.lower()}"

                with _get_db() as conn:
                    batch_id = log_upload(conn, file.filename, detected_type, sector,
                                          len(daily), len(daily) - len(errors), len(errors),
                                          min(dates) if dates else None, max(dates) if dates else None,
                                          errors[:50] if errors else None)
                    gen_id_map = {}
                    for gen in gens:
                        upsert_site(conn, gen["site_id"], gen["site_name"], sector, gen["site_type"],
                                    cost_center_code=gen.get("cost_center_code"),
                                    business_sector=gen.get("business_sector"),
                                    company=gen.get("company"),
                                    site_code=gen.get("site_code"))
                        gen_id = upsert_generator(conn, gen["site_id"], gen["model_name"],
                                                   gen["model_name_raw"], gen["power_kva"],
                                                   gen["consumption_per_hour"], gen["fuel_type"], gen["supplier"])
                        gen_id_map[(gen["site_id"], gen["model_name_raw"])] = gen_id

                    sites_dates = set()
                    for row in daily:
                        gen_key = (row["site_id"], row["model_name_raw"])
                        gen_id = gen_id_map.get(gen_key)
                        if gen_id:
                            upsert_daily_operation(conn, gen_id, row["site_id"], row["date"],
                                                    row["gen_run_hr"], row["daily_used_liters"],
                                                    row["spare_tank_balance"], row["blackout_hr"],
                                                    source="api", upload_batch_id=batch_id)
                            sites_dates.add((row["site_id"], row["date"]))
                    for sid, dt in sites_dates:
                        refresh_site_summary(conn, sid, dt)

                    # Map sales site_id: since site_id = cost_center_code,
                    # sales_site_name (CostCenter) directly matches site_id
                    conn.execute("""
                        UPDATE daily_sales SET site_id = sales_site_name
                        WHERE site_id IS NULL AND EXISTS (
                            SELECT 1 FROM sites s WHERE s.site_id = daily_sales.sales_site_name
                        )
                    """)
                    conn.execute("""
                        UPDATE hourly_sales SET site_id = sales_site_name
                        WHERE site_id IS NULL AND EXISTS (
                            SELECT 1 FROM sites s WHERE s.site_id = hourly_sales.sales_site_name
                        )
                    """)
                    # Fallback: match via cost_center_code
                    conn.execute("""
                        UPDATE daily_sales SET site_id = (
                            SELECT s.site_id FROM sites s
                            WHERE s.cost_center_code = daily_sales.sales_site_name
                            LIMIT 1
                        )
                        WHERE site_id IS NULL AND EXISTS (
                            SELECT 1 FROM sites s WHERE s.cost_center_code = daily_sales.sales_site_name
                        )
                    """)
                    conn.execute("""
                        UPDATE hourly_sales SET site_id = (
                            SELECT s.site_id FROM sites s
                            WHERE s.cost_center_code = hourly_sales.sales_site_name
                            LIMIT 1
                        )
                        WHERE site_id IS NULL AND EXISTS (
                            SELECT 1 FROM sites s WHERE s.cost_center_code = hourly_sales.sales_site_name
                        )
                    """)

                    # Enrich sites with store master data (address, lat/long, etc.)
                    try:
                        enrich_sites_from_store_master(conn)
                    except Exception:
                        pass

                # ─── Build validation comparison: Excel vs Pipeline vs DB ───
                try:
                    import openpyxl as _xl
                    from parsers.base_parser import parse_date_from_cell as _pdc, clean_numeric as _cn
                    from collections import defaultdict as _dd

                    # 1. Excel totals from total row
                    _wb = _xl.load_workbook(tmp_path, data_only=True)
                    _ws = _wb[sector] if sector in _wb.sheetnames else _wb[_wb.sheetnames[0]]
                    _total_row = None
                    for _r in range(1, 15):
                        for _c in range(1, 8):
                            _v = _ws.cell(row=_r, column=_c).value
                            if _v and "total" in str(_v).lower():
                                _total_row = _r; break
                        if _total_row: break

                    _cd = parsed.get("columns_detected", {})
                    _dsub = _cd.get("date_sub_offsets", {})
                    _og, _of, _ot, _ob = _dsub.get("gen_run_hr",0), _dsub.get("daily_used_liters",1), _dsub.get("spare_tank_balance",2), _dsub.get("blackout_hr",3)
                    _dcols = []
                    for _ci in range(1, _ws.max_column + 1):
                        _ds = _pdc(_ws.cell(row=2, column=_ci).value)
                        if _ds: _dcols.append((_ds, _ci))

                    _excel = {}
                    if _total_row:
                        for _ds, _dc in _dcols:
                            _excel[_ds] = {
                                "gen_hr": _cn(_ws.cell(row=_total_row, column=_dc+_og).value) or 0,
                                "fuel": _cn(_ws.cell(row=_total_row, column=_dc+_of).value) or 0,
                                "tank": _cn(_ws.cell(row=_total_row, column=_dc+_ot).value) or 0,
                                "blackout": _cn(_ws.cell(row=_total_row, column=_dc+_ob).value) or 0,
                            }
                    _wb.close()

                    # Save Excel totals to cache for DATA_QUALITY tab
                    with _get_db() as _cache_conn:
                        for _ds, _ev in _excel.items():
                            _cache_conn.execute("""
                                INSERT OR REPLACE INTO excel_validation_cache
                                (sector_id, date, gen_hr, fuel, tank, blackout)
                                VALUES (?, ?, ?, ?, ?, ?)
                            """, (sector, _ds, _ev["gen_hr"], _ev["fuel"], _ev["tank"], _ev["blackout"]))

                    # 2. Parser totals
                    _ptot = _dd(lambda: {"gen_hr":0,"fuel":0,"tank":0,"blackout":0,"sites":set()})
                    for _row in daily:
                        _d = _row["date"]
                        _ptot[_d]["gen_hr"] += _row["gen_run_hr"] or 0
                        _ptot[_d]["fuel"] += _row["daily_used_liters"] or 0
                        _ptot[_d]["tank"] += _row["spare_tank_balance"] or 0
                        _ptot[_d]["blackout"] += _row["blackout_hr"] or 0
                        _ptot[_d]["sites"].add(_row["site_id"])

                    # 3. DB totals
                    import pandas as _pd
                    _dbdf = _pd.read_sql_query("""
                        SELECT dss.date, SUM(dss.total_gen_run_hr) as gen_hr, SUM(dss.total_daily_used) as fuel,
                               SUM(dss.spare_tank_balance) as tank, SUM(dss.blackout_hr) as blackout,
                               COUNT(DISTINCT dss.site_id) as sites
                        FROM daily_site_summary dss JOIN sites s ON s.site_id=dss.site_id
                        WHERE s.sector_id=? GROUP BY dss.date ORDER BY dss.date
                    """, conn, params=[sector])
                    _dbtot = {}
                    for _, _r in _dbdf.iterrows():
                        _dbtot[_r["date"]] = {"gen_hr":_r["gen_hr"] or 0,"fuel":_r["fuel"] or 0,"tank":_r["tank"] or 0,"blackout":_r["blackout"] or 0,"sites":int(_r["sites"])}

                    # 4. Build comparison rows
                    validation = []
                    _all_dates = sorted(set(list(_excel.keys()) + list(_ptot.keys()) + list(_dbtot.keys())))
                    for _d in _all_dates:
                        _e = _excel.get(_d, {"gen_hr":0,"fuel":0,"tank":0,"blackout":0})
                        _p = _ptot.get(_d, {"gen_hr":0,"fuel":0,"tank":0,"blackout":0,"sites":set()})
                        _db = _dbtot.get(_d, {"gen_hr":0,"fuel":0,"tank":0,"blackout":0,"sites":0})
                        _pass = all(abs(_e[k] - _db[k]) < 2 for k in ["gen_hr","fuel","tank","blackout"])
                        validation.append({
                            "date": _d,
                            "sites": len(_p["sites"]) if isinstance(_p["sites"], set) else 0,
                            "excel": {k: round(_e[k], 1) for k in ["gen_hr","fuel","tank","blackout"]},
                            "pipeline": {k: round(_p[k], 1) for k in ["gen_hr","fuel","tank","blackout"]},
                            "db": {k: round(_db[k], 1) for k in ["gen_hr","fuel","tank","blackout"]},
                            "pass": _pass,
                        })
                except Exception:
                    validation = []

                record_count = len(daily)

        # ─── COMBO SALES ───────────────────────────────────────
        elif any(s in sheets_lower for s in ["daily sales", "daily_sales", "hourly sales", "hourly_sales"]):
            from parsers.sales_parser import parse_daily_sales_file, parse_hourly_sales_file
            total = 0
            is_large = False

            # Daily sales
            for sn in sheets:
                if "daily" in sn.lower() and "sales" in sn.lower():
                    try:
                        result = parse_daily_sales_file(tmp_path, sheet_name=sn)
                        records = result.get("records", [])
                        is_large = len(records) > 10000

                        with _get_db() as conn:
                            ftype = "sales_reference" if is_large else "daily_sales"
                            batch_id = log_upload(conn, file.filename, ftype, None,
                                                  len(records), len(records), 0,
                                                  result.get("date_range", [None, None])[0],
                                                  result.get("date_range", [None, None])[1], None)

                            if is_large:
                                # Chunked batch insert for large files (10K per transaction)
                                CHUNK = 10000
                                for start in range(0, len(records), CHUNK):
                                    chunk = records[start:start + CHUNK]
                                    conn.executemany("""
                                        INSERT OR REPLACE INTO daily_sales
                                        (sales_site_name, sector_id, date, brand, sales_amt, margin, source, upload_batch_id)
                                        VALUES (?, ?, ?, ?, ?, ?, 'api', ?)
                                    """, [(r["sales_site_name"], r.get("sector_id"), r["date"],
                                           r["brand"], r["sales_amt"], r["margin"], batch_id) for r in chunk])
                            else:
                                for r in records:
                                    upsert_daily_sale(conn, r["sales_site_name"], r.get("sector_id"),
                                                      r["date"], r["brand"], r["sales_amt"], r["margin"],
                                                      source="api", upload_batch_id=batch_id)
                        total += len(records)
                    except Exception as e:
                        import logging; logging.error(f"Daily sales parse failed for sheet '{sn}': {e}")
                        import traceback; traceback.print_exc()

            # Hourly sales
            for sn in sheets:
                if "hourly" in sn.lower() and "sales" in sn.lower():
                    try:
                        result = parse_hourly_sales_file(tmp_path, sheet_name=sn)
                        records = result.get("records", [])

                        with _get_db() as conn:
                            ftype = "hourly_reference" if is_large else "hourly_sales"
                            log_upload(conn, file.filename, ftype, None,
                                       len(records), len(records), 0, None, None, None)

                            if is_large:
                                CHUNK = 10000
                                for start in range(0, len(records), CHUNK):
                                    chunk = records[start:start + CHUNK]
                                    conn.executemany("""
                                        INSERT OR REPLACE INTO hourly_sales
                                        (sales_site_name, sector_id, date, hour, brand, sales_amt, trans_cnt, source)
                                        VALUES (?, ?, ?, ?, ?, ?, ?, 'api')
                                    """, [(r["sales_site_name"], r.get("sector_id"), r["date"],
                                           r["hour"], r["brand"], r["sales_amt"], r["trans_cnt"]) for r in chunk])
                            else:
                                for r in records:
                                    upsert_hourly_sale(conn, r["sales_site_name"], r.get("sector_id"),
                                                       r["date"], r["hour"], r["brand"],
                                                       r["sales_amt"], r["trans_cnt"], source="api")
                        total += len(records)
                    except Exception as e:
                        import logging; logging.error(f"Hourly sales parse failed for sheet '{sn}': {e}")
                        import traceback; traceback.print_exc()

            # ─── STORE MASTER sheets (auto-detect from combo file) ───
            storemaster_sheets = []
            for sn in sheets:
                sn_lower = sn.lower().strip()
                if sn_lower in ("store master", "store_master", "storemaster"):
                    storemaster_sheets.append(sn)
                elif "all" in sn_lower and "sites" in sn_lower:
                    storemaster_sheets.append(sn)
                elif sn_lower == "cp" and any("cmhl" in s.lower() or "sites" in s.lower() for s in sheets):
                    # CP sheet in a combo file (not a standalone blackout file)
                    storemaster_sheets.append(sn)

            if storemaster_sheets:
                try:
                    sm_total = 0
                    for sn in storemaster_sheets:
                        try:
                            import pandas as _pd_sm
                            _df_sm = _pd_sm.read_excel(tmp_path, sheet_name=sn)

                            with _get_db() as conn:
                                # Detect columns
                                cols_lower = {str(c).lower().replace(" ", "").replace("_", ""): c for c in _df_sm.columns}
                                gold_col = next((cols_lower[k] for k in cols_lower if "goldcode" in k), None)
                                pos_col = next((cols_lower[k] for k in cols_lower if "poscode" in k), None)
                                ccd_col = next((cols_lower[k] for k in cols_lower if "costcenterdescription" in k), None)
                                ccn_col = next((cols_lower[k] for k in cols_lower if "costcentername" in k), None)
                                cc_col = next((cols_lower[k] for k in cols_lower if k == "costcenter"), None)
                                if not cc_col:
                                    cc_col = next((cols_lower[k] for k in cols_lower if "costcenter" in k and "name" not in k and "description" not in k), None)
                                seg_col = next((cols_lower[k] for k in cols_lower if "segmentname" in k), None)
                                sector_col = next((cols_lower[k] for k in cols_lower if k == "sector"), None)
                                company_col = next((cols_lower[k] for k in cols_lower if "companycode" in k), None)
                                legal_col = next((cols_lower[k] for k in cols_lower if "legalentity" in k), None)
                                channel_col = next((cols_lower[k] for k in cols_lower if "channel" in k), None)
                                state_col = next((cols_lower[k] for k in cols_lower if "addressstate" in k), None)
                                township_col = next((cols_lower[k] for k in cols_lower if "addresstownship" in k), None)
                                lat_col = next((cols_lower[k] for k in cols_lower if "latitude" in k), None)
                                lon_col = next((cols_lower[k] for k in cols_lower if "longitude" in k), None)
                                size_col = next((cols_lower[k] for k in cols_lower if "storesize" in k), None)
                                open_col = next((cols_lower[k] for k in cols_lower if "opendate" in k), None)
                                close_col = next((cols_lower[k] for k in cols_lower if "closeddate" in k or "closedate" in k), None)
                                cpid_col = next((cols_lower[k] for k in cols_lower if "cpcenterid" in k), None)

                                if gold_col:
                                    for _, row in _df_sm.iterrows():
                                        gc = row.get(gold_col)
                                        if pd.isna(gc):
                                            continue
                                        gc_str = str(int(gc)) if isinstance(gc, float) else str(gc).strip()
                                        cc = row.get(cc_col) if cc_col else None
                                        cc_str = str(int(cc)) if isinstance(cc, float) and cc and not pd.isna(cc) else (str(cc).strip() if cc and not pd.isna(cc) else None)
                                        ccn = str(row.get(ccn_col, '')).strip() if ccn_col and not pd.isna(row.get(ccn_col)) else ''
                                        ccd = str(row.get(ccd_col, '')).strip() if ccd_col and not pd.isna(row.get(ccd_col)) else None
                                        seg = str(row.get(seg_col, '')).strip() if seg_col and not pd.isna(row.get(seg_col)) else ''
                                        sec = str(row.get(sector_col, '')).strip() if sector_col and not pd.isna(row.get(sector_col)) else ''
                                        comp = str(row.get(company_col, '')).strip() if company_col and not pd.isna(row.get(company_col)) else None
                                        legal = str(row.get(legal_col, '')).strip() if legal_col and not pd.isna(row.get(legal_col)) else None
                                        cpid = str(row.get(cpid_col, '')).strip() if cpid_col and not pd.isna(row.get(cpid_col)) else None
                                        # Map sector name to sector_id
                                        sec_map = {"retail": "CMHL", "property": "CP", "f&b": "CFC", "distribution": "PG"}
                                        sector_id = sec_map.get(sec.lower(), '') if sec else ''
                                        if not sector_id:
                                            from config.settings import SEGMENT_SECTOR_MAP
                                            sector_id = SEGMENT_SECTOR_MAP.get(seg, '')

                                        lat = row.get(lat_col) if lat_col and not pd.isna(row.get(lat_col, None)) else None
                                        lon = row.get(lon_col) if lon_col and not pd.isna(row.get(lon_col, None)) else None

                                        conn.execute("""
                                            INSERT OR REPLACE INTO store_master
                                            (gold_code, pos_code, store_name, cost_center_code, cost_center_name,
                                             cost_center_description, segment_name, company_code, legal_entity,
                                             channel, address_state, address_township,
                                             latitude, longitude, store_size, open_date, closed_date,
                                             sector_id, cp_center_id)
                                            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                                        """, (
                                            gc_str,
                                            str(row.get(pos_col, '')).strip() if pos_col and not pd.isna(row.get(pos_col)) else '',
                                            ccn or gc_str,
                                            cc_str,
                                            ccn,
                                            ccd,
                                            seg,
                                            comp,
                                            legal,
                                            str(row.get(channel_col, '')).strip() if channel_col and not pd.isna(row.get(channel_col)) else None,
                                            str(row.get(state_col, '')).strip() if state_col and not pd.isna(row.get(state_col)) else None,
                                            str(row.get(township_col, '')).strip() if township_col and not pd.isna(row.get(township_col)) else None,
                                            _parse_coord(lat),
                                            _parse_coord(lon),
                                            str(row.get(size_col, '')).strip() if size_col and not pd.isna(row.get(size_col)) else None,
                                            str(row.get(open_col, '')) if open_col and not pd.isna(row.get(open_col)) else None,
                                            str(row.get(close_col, '')) if close_col and not pd.isna(row.get(close_col)) else None,
                                            sector_id or None,
                                            cpid or None,
                                        ))
                                        sm_total += 1

                                log_upload(conn, file.filename, f"store_master_{sn}", None,
                                           sm_total, sm_total, 0, None, None, None)
                        except Exception as e:
                            import traceback; traceback.print_exc()
                    total += sm_total
                    # Enrich sites with store master data
                    try:
                        with _get_db() as conn:
                            enrich_sites_from_store_master(conn)
                    except Exception:
                        pass
                except Exception:
                    pass

            # ─── MAPPING sheet ─────────────────────────────────────
            for sn in sheets:
                if sn.lower().strip() == "mapping":
                    try:
                        import pandas as _pd_map
                        _df_map = _pd_map.read_excel(tmp_path, sheet_name=sn)
                        with _get_db() as conn:
                            for _, row in _df_map.iterrows():
                                manual = row.get(_df_map.columns[0])
                                sap_cc = row.get(_df_map.columns[1]) if len(_df_map.columns) > 1 else None
                                site_code = row.get(_df_map.columns[2]) if len(_df_map.columns) > 2 else None
                                if pd.isna(manual) or pd.isna(site_code):
                                    continue
                                # Insert into site_sales_map
                                cc_str = str(int(sap_cc)) if isinstance(sap_cc, float) and not pd.isna(sap_cc) else (str(sap_cc).strip() if sap_cc and not pd.isna(sap_cc) else '')
                                conn.execute("""
                                    INSERT OR REPLACE INTO site_sales_map
                                    (sales_site_name, site_id, sector_id, match_method)
                                    VALUES (?, ?, (SELECT sector_id FROM sites WHERE site_id = ?), 'manual')
                                """, (str(manual).strip(), str(site_code).strip(), str(site_code).strip()))
                            log_upload(conn, file.filename, "mapping", None,
                                       len(_df_map), len(_df_map), 0, None, None, None)
                    except Exception as e:
                        import logging; logging.error(f"Mapping sheet parse failed: {e}")

            detected_type = "sales_reference" if is_large else "combo_sales"
            record_count = total

        # ─── DIESEL EXPENSE LY ─────────────────────────────────
        else:
            try:
                import pandas as _pd
                _df = _pd.read_excel(tmp_path, sheet_name=0, nrows=5)
                all_text = ' '.join(str(c).lower() for c in _df.columns)

                if "store contribution" in all_text or ("store" in all_text and "center" in all_text and "cost center code" in all_text):
                    # Store-to-center allocation mapping
                    import pandas as _pd_alloc
                    _df_alloc = _pd_alloc.read_excel(tmp_path, sheet_name=0)
                    detected_type = "store_allocation"
                    alloc_count = 0

                    with _get_db() as conn:
                        for _, row in _df_alloc.iterrows():
                            store_cc = row.iloc[2] if len(row) > 2 else None  # Store's Cost Center Code
                            center_cc = row.iloc[5] if len(row) > 5 else None  # Center's Cost Center Code
                            alloc_pct = row.iloc[6] if len(row) > 6 else None  # Store Contribution %
                            if pd.isna(store_cc) or pd.isna(center_cc) or pd.isna(alloc_pct):
                                continue
                            store_cc = str(int(store_cc)) if isinstance(store_cc, float) else str(store_cc).strip()
                            center_cc = str(int(center_cc)) if isinstance(center_cc, float) else str(center_cc).strip()
                            store_name = str(row.iloc[1]).strip() if not pd.isna(row.iloc[1]) else ''
                            center_name = str(row.iloc[4]).strip() if not pd.isna(row.iloc[4]) else ''
                            remark = str(row.iloc[7]).strip() if len(row) > 7 and not pd.isna(row.iloc[7]) else ''
                            pct = float(alloc_pct) if isinstance(alloc_pct, (int, float)) else 0
                            # Convert from decimal (0.50) to percentage (50.0) if needed
                            if pct < 1:
                                pct = round(pct * 100, 2)

                            conn.execute("""
                                INSERT OR REPLACE INTO store_center_allocation
                                (store_cost_center, store_name, center_cost_center, center_name, allocation_pct, remark)
                                VALUES (?, ?, ?, ?, ?, ?)
                            """, (store_cc, store_name, center_cc, center_name, pct, remark))
                            alloc_count += 1
                        log_upload(conn, file.filename, "store_allocation", None,
                                   alloc_count, alloc_count, 0, None, None, None)
                    record_count = alloc_count

                elif any(kw in all_text for kw in ["daily average diesel", "daily normal diesel", "diesel expense", "store exp"]):
                    from parsers.diesel_expense_parser import parse_diesel_expense_file
                    parsed = parse_diesel_expense_file(tmp_path)
                    records = parsed.get("records", []) if isinstance(parsed, dict) else []
                    detected_type = "diesel_expense_ly"

                    with _get_db() as conn:
                        log_upload(conn, file.filename, "diesel_expense_ly", None,
                                   len(records), len(records), 0, None, None, None)
                        for r in records:
                            upsert_diesel_expense_ly(conn, r["cost_center_code"], r["sector_id"],
                                                      r["cost_center_name"],
                                                      r["yearly_expense_mmk_mil"],
                                                      r["daily_avg_expense_mmk"],
                                                      r["pct_on_sales"])
                    record_count = len(records)

                elif any(kw in all_text for kw in ["sales_amt", "sales_date", "gold_code"]):
                    from parsers.sales_parser import parse_daily_sales_file
                    parsed = parse_daily_sales_file(tmp_path)
                    records = parsed.get("records", [])
                    detected_type = "daily_sales"

                    with _get_db() as conn:
                        log_upload(conn, file.filename, "daily_sales", None,
                                   len(records), len(records), 0, None, None, None)
                        for r in records:
                            upsert_daily_sale(conn, r["sales_site_name"], r.get("sector_id"),
                                              r["date"], r["brand"], r["sales_amt"], r["margin"],
                                              source="api")
                    record_count = len(records)
            except Exception as e:
                import logging; logging.error(f"Diesel expense/fallback parse failed: {e}")
                import traceback; traceback.print_exc()

        # Map sales site_id: since site_id = cost_center_code, direct match
        if detected_type in ("combo_sales", "sales_reference", "daily_sales", "hourly_sales", "hourly_reference"):
            try:
                with _get_db() as conn:
                    # 1. Direct match: sales_site_name == site_id
                    conn.execute("""
                        UPDATE daily_sales SET site_id = sales_site_name
                        WHERE site_id IS NULL AND EXISTS (
                            SELECT 1 FROM sites s WHERE s.site_id = daily_sales.sales_site_name
                        )
                    """)
                    conn.execute("""
                        UPDATE hourly_sales SET site_id = sales_site_name
                        WHERE site_id IS NULL AND EXISTS (
                            SELECT 1 FROM sites s WHERE s.site_id = hourly_sales.sales_site_name
                        )
                    """)
                    # 2. Fallback: sales_site_name matches cost_center_code in sites
                    conn.execute("""
                        UPDATE daily_sales SET site_id = (
                            SELECT s.site_id FROM sites s
                            WHERE s.cost_center_code = daily_sales.sales_site_name
                            LIMIT 1
                        )
                        WHERE site_id IS NULL AND EXISTS (
                            SELECT 1 FROM sites s WHERE s.cost_center_code = daily_sales.sales_site_name
                        )
                    """)
                    conn.execute("""
                        UPDATE hourly_sales SET site_id = (
                            SELECT s.site_id FROM sites s
                            WHERE s.cost_center_code = hourly_sales.sales_site_name
                            LIMIT 1
                        )
                        WHERE site_id IS NULL AND EXISTS (
                            SELECT 1 FROM sites s WHERE s.cost_center_code = hourly_sales.sales_site_name
                        )
                    """)
                    # 3. Fallback: match via site_sales_map (populated from mapping sheet)
                    conn.execute("""
                        UPDATE daily_sales SET site_id = (
                            SELECT m.site_id FROM site_sales_map m
                            WHERE m.sales_site_name = daily_sales.sales_site_name
                            LIMIT 1
                        )
                        WHERE site_id IS NULL AND EXISTS (
                            SELECT 1 FROM site_sales_map m WHERE m.sales_site_name = daily_sales.sales_site_name
                        )
                    """)
                    conn.execute("""
                        UPDATE hourly_sales SET site_id = (
                            SELECT m.site_id FROM site_sales_map m
                            WHERE m.sales_site_name = hourly_sales.sales_site_name
                            LIMIT 1
                        )
                        WHERE site_id IS NULL AND EXISTS (
                            SELECT 1 FROM site_sales_map m WHERE m.sales_site_name = hourly_sales.sales_site_name
                        )
                    """)
            except Exception as e:
                import logging; logging.error(f"Sales site_id mapping failed: {e}")

        # Run alert checks after blackout/fuel uploads
        if detected_type and detected_type.startswith("blackout") or detected_type == "fuel_price":
            try:
                from alerts.alert_engine import run_all_checks
                run_all_checks()
            except Exception:
                pass

        try:
            with get_db() as _log_conn:
                log_activity(_log_conn, user["id"], user["username"], "UPLOAD", "DATA", f"Uploaded {file.filename} — {record_count} records ({detected_type})", {"filename": file.filename, "records": record_count, "type": detected_type})
        except Exception:
            pass

        return {
            "filename": file.filename,
            "sheets": sheets,
            "records": record_count,
            "type": detected_type or "unknown",
            "validation": (validation if 'validation' in dir() else []) + col_warnings,
        }
    except Exception as e:
        raise HTTPException(400, f"Parse error: {e}")
    finally:
        os.unlink(tmp_path)


# ─── 1.7 Economics ───────────────────────────────────────
@router.get("/economics")
def economics(
    date_from: Optional[str] = None, date_to: Optional[str] = None,
    sector: Optional[str] = None, site_type: Optional[str] = None,
    user: dict = Depends(get_current_user),
):
    df = get_store_economics(date_from=date_from, date_to=date_to)
    if df.empty:
        return []
    if sector:
        df = df[df["sector_id"] == sector]
    if site_type and site_type != "All":
        df = df[df["site_type"] == site_type]
    return df.fillna("").to_dict(orient="records")


# ─── 1.8 Daily Summary ──────────────────────────────────
@router.get("/daily-summary")
def daily_summary(
    date_from: Optional[str] = None, date_to: Optional[str] = None,
    sector: Optional[str] = None, site_id: Optional[str] = None,
    site_type: Optional[str] = None,
    user: dict = Depends(get_current_user),
):
    with get_db() as conn:
        dsql, dp = _dsql("dss.date", date_from, date_to)
        q = f"""SELECT dss.*, s.sector_id, s.site_type,
                       s.cost_center_code, s.region as site_code, s.segment_name,
                       s.cost_center_description, s.address_state, s.store_size,
                       s.latitude, s.longitude
                FROM daily_site_summary dss
                JOIN sites s ON dss.site_id = s.site_id WHERE 1=1{dsql}"""
        if sector:
            q += " AND s.sector_id = ?"; dp.append(sector)
        if site_id:
            q += " AND dss.site_id = ?"; dp.append(site_id)
        if site_type and site_type != 'All':
            q += " AND s.site_type = ?"; dp.append(site_type)
        q += " ORDER BY dss.date DESC LIMIT 5000"
        df = pd.read_sql_query(q, conn, params=dp)
    return df.fillna("").to_dict(orient="records")


# ─── 1.9 Fuel Prices ────────────────────────────────────
@router.get("/fuel-prices")
def fuel_prices(
    date_from: Optional[str] = None, date_to: Optional[str] = None,
    sector: Optional[str] = None,
    user: dict = Depends(get_current_user),
):
    with get_db() as conn:
        dsql, dp = _dsql("date", date_from, date_to)
        q = f"SELECT * FROM fuel_purchases WHERE 1=1{dsql}"
        if sector:
            q += " AND sector_id = ?"; dp.append(sector)
        q += " ORDER BY date DESC LIMIT 2000"
        df = pd.read_sql_query(q, conn, params=dp)
    return df.fillna("").to_dict(orient="records")


# ─── 1.10 Sales ──────────────────────────────────────────
@router.get("/sales")
def sales(
    date_from: Optional[str] = None, date_to: Optional[str] = None,
    site_id: Optional[str] = None,
    user: dict = Depends(get_current_user),
):
    with get_db() as conn:
        dsql, dp = _dsql("date", date_from, date_to)
        q = f"SELECT * FROM daily_sales WHERE 1=1{dsql}"
        if site_id:
            q += " AND site_id = ?"; dp.append(site_id)
        q += " ORDER BY date DESC LIMIT 5000"
        df = pd.read_sql_query(q, conn, params=dp)
    return df.fillna("").to_dict(orient="records")


# ─── 1.11 Blackout ───────────────────────────────────────
@router.get("/blackout")
def blackout(
    date_from: Optional[str] = None, date_to: Optional[str] = None,
    sector: Optional[str] = None,
    site_type: Optional[str] = None,
    user: dict = Depends(get_current_user),
):
    with get_db() as conn:
        dsql, dp = _dsql("do.date", date_from, date_to)
        q = f"""SELECT do.date, do.site_id, do.blackout_hr, do.gen_run_hr,
                       do.daily_used_liters, s.sector_id
                FROM daily_operations do JOIN sites s ON do.site_id = s.site_id
                WHERE do.blackout_hr IS NOT NULL{dsql}"""
        if sector:
            q += " AND s.sector_id = ?"; dp.append(sector)
        if site_type and site_type != 'All':
            q += " AND s.site_type = ?"; dp.append(site_type)
        q += " ORDER BY do.date DESC LIMIT 5000"
        df = pd.read_sql_query(q, conn, params=dp)
    return df.fillna("").to_dict(orient="records")


# ─── 1.12 Generators ────────────────────────────────────
@router.get("/generators")
def generators(
    sector: Optional[str] = None, site_id: Optional[str] = None,
    site_type: Optional[str] = None,
    user: dict = Depends(get_current_user),
):
    with get_db() as conn:
        q = """SELECT g.*, s.sector_id, s.site_type FROM generators g
               JOIN sites s ON g.site_id = s.site_id WHERE g.is_active = 1"""
        dp = []
        if sector:
            q += " AND s.sector_id = ?"; dp.append(sector)
        if site_id:
            q += " AND g.site_id = ?"; dp.append(site_id)
        if site_type and site_type != 'All':
            q += " AND s.site_type = ?"; dp.append(site_type)
        q += " ORDER BY g.site_id"
        df = pd.read_sql_query(q, conn, params=dp)
    return df.fillna("").to_dict(orient="records")


# ─── 1.13 Site Detail ────────────────────────────────────
@router.get("/site/{site_id}")
def site_detail(
    site_id: str,
    date_from: Optional[str] = None, date_to: Optional[str] = None,
    user: dict = Depends(get_current_user),
):
    econ = get_store_economics(date_from=date_from, date_to=date_to)
    site_row = econ[econ["site_id"] == site_id]
    if site_row.empty:
        raise HTTPException(404, f"Site {site_id} not found")

    gen_detail = get_generator_detail(site_id, date_from=date_from, date_to=date_to)
    trends = get_trends(site_id, date_from=date_from, date_to=date_to)

    return {
        "economics": site_row.fillna("").to_dict(orient="records")[0],
        "generators": gen_detail.fillna("").to_dict(orient="records") if not gen_detail.empty else [],
        "trends": {k: v.fillna("").to_dict(orient="records") if isinstance(v, pd.DataFrame) and not v.empty else [] for k, v in trends.items()} if trends else {},
    }
