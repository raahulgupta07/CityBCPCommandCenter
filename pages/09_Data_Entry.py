"""
Page 9: Data Entry — Upload all Excel files at once, system auto-detects type
"""
import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent.parent))

import streamlit as st
import streamlit_shadcn_ui as ui
import pandas as pd
import tempfile
import os
from datetime import date
from utils.database import (
    get_db, upsert_site, upsert_generator, upsert_daily_operation,
    insert_fuel_purchase, refresh_site_summary, log_upload,
    upsert_store_master, upsert_daily_sale, upsert_hourly_sale,
    upsert_diesel_expense_ly,
)
from parsers.blackout_parser import parse_blackout_file
from parsers.fuel_price_parser import parse_fuel_price_file
from parsers.sales_parser import parse_daily_sales_file, parse_hourly_sales_file
from parsers.storemaster_parser import parse_storemaster_file
from parsers.diesel_expense_parser import parse_diesel_expense_file
from models.energy_cost import auto_map_sites
from alerts.alert_engine import run_all_checks, get_active_alerts
from utils.email_sender import is_email_configured, send_alert_email
from config.settings import SECTORS
from utils.page_header import render_page_header
from utils.auth import require_login, require_role, render_sidebar_user, has_permission

st.set_page_config(page_title="Data Entry", page_icon="📤", layout="wide")
require_login()
require_role("admin")  # Only admin and super_admin can upload data
render_sidebar_user()

render_page_header("📤", "Data Entry", "Upload Excel files, browse raw data, and manage upload history")

ui.alert(
    title="📤 Upload All Your Excel Files At Once",
    description="Drop all files together (up to 8). Auto-detects: blackout, fuel price, daily sales, hourly sales, store master, and site location mapping. File names don't matter.",
    key="alert_entry",
)

selected = ui.tabs(
    options=["📁 Upload Files", "📊 Store Summary", "🗄️ Raw Data", "🧹 How AI Cleans Data", "📋 History"],
    default_value="📁 Upload Files",
    key="entry_tabs",
)


# ═══════════════════════════════════════════════════════════════════════════
# AUTO-DETECT: Read sheet names to determine file type
# ═══════════════════════════════════════════════════════════════════════════
def detect_file_type(filepath):
    """
    Detect file type by reading sheet names inside the Excel file.
    Returns (file_type, sector_id) or (None, None) if unknown.
    """
    import openpyxl
    try:
        wb = openpyxl.load_workbook(str(filepath), read_only=True)
        sheets = [s.lower().strip() for s in wb.sheetnames]
        wb.close()

        # Site location mapping file: check for Manual Data + SAP Cost Center columns
        try:
            import pandas as _pd
            _df = _pd.read_excel(str(filepath), sheet_name=0, nrows=2)
            if "Manual Data" in _df.columns and "SAP Cost Center" in _df.columns:
                return "site_mapping", None
            # Diesel expense LY file: has "Daily Average Diesel Exp" column
            if any("daily average diesel" in str(c).lower() for c in _df.columns):
                return "diesel_expense_ly", None
        except Exception:
            pass

        # Combo sales file: has daily + hourly + store master in one file
        has_daily = any("daily sales" in s or "daily_sales" in s for s in sheets)
        has_hourly = any("hourly sales" in s or "hourly_sales" in s for s in sheets)
        has_master = any("store master" in s or "storemaster" in s for s in sheets)
        has_mapping = any(s == "mapping" for s in sheets)

        if has_daily and has_hourly:
            return "combo_sales", None  # multi-sheet file

        # Single-purpose sales files
        if has_daily:
            return "daily_sales", None
        if has_hourly:
            return "hourly_sales", None
        if has_master:
            return "store_master", None

        # Fuel price: has multiple sector sheets (CMHL, CP, CFC, PG)
        if len(sheets) >= 3 and any(s in sheets for s in ["cmhl", "cp", "cfc", "pg"]):
            return "fuel_price", None

        # Blackout files: single sheet named after sector
        if "cp" in sheets:
            return "blackout_cp", "CP"
        if "cmhl" in sheets:
            return "blackout_cmhl", "CMHL"
        if "cfc" in sheets:
            return "blackout_cfc", "CFC"

        # Fallback: check first sheet name
        first = sheets[0] if sheets else ""
        if "cp" in first:
            return "blackout_cp", "CP"
        if "cmhl" in first or "cm" in first:
            return "blackout_cmhl", "CMHL"
        if "cfc" in first:
            return "blackout_cfc", "CFC"

    except Exception:
        pass
    return None, None


def process_file(tmp_path, filename):
    """Process a single uploaded file. Returns dict with results."""
    file_type, sector_id = detect_file_type(tmp_path)

    if file_type is None:
        return {"filename": filename, "status": "❌ Unknown", "type": "?",
                "message": "Could not detect file type from sheet names", "rows": 0}

    # ─── Combo Sales File (daily + hourly + store master + mapping) ───
    if file_type == "combo_sales":
        return _process_combo_sales(tmp_path, filename)

    # ─── Diesel Expense LY ──────────────────────────────────────────────
    if file_type == "diesel_expense_ly":
        result = parse_diesel_expense_file(tmp_path)
        records = result["records"]
        if not records:
            return {"filename": filename, "status": "⚠️ Empty", "type": "Diesel Expense LY",
                    "message": "No records found", "rows": 0}

        with get_db() as conn:
            batch_id = log_upload(conn, filename, "diesel_expense_ly", None,
                                  len(records), len(records), len(result["errors"]),
                                  None, None, result["errors"][:50] if result["errors"] else None)
            for r in records:
                upsert_diesel_expense_ly(
                    conn, r["cost_center_code"], r["sector_id"], r["cost_center_name"],
                    r["yearly_expense_mmk_mil"], r["daily_avg_expense_mmk"], r["pct_on_sales"],
                )

        sector_info = ", ".join(f"{k}: {v}" for k, v in result["sector_counts"].items())
        return {"filename": filename, "status": "✅ Imported", "type": "Diesel Expense LY",
                "message": f"{len(records)} stores ({sector_info})",
                "rows": len(records), "data": pd.DataFrame(records[:20])}

    # ─── Site Location Mapping ─────────────────────────────────────────
    if file_type == "site_mapping":
        matched, unmatched_count, details, unmatched_info = auto_map_sites(mapping_file=tmp_path)
        if isinstance(unmatched_info, str):
            # Error message
            return {"filename": filename, "status": "❌ Error", "type": "Site Mapping",
                    "message": unmatched_info, "rows": 0}
        return {"filename": filename, "status": "✅ Imported", "type": "Site Mapping",
                "message": f"{matched} BCP sites mapped to sales data, {unmatched_count} unmatched",
                "rows": matched,
                "data": pd.DataFrame(details) if details else pd.DataFrame(),
                "errors": [f"{u['bcp']} → {u.get('sap','—')}: {u['reason']}" for u in (unmatched_info or [])]}

    # ─── Store Master ────────────────────────────────────────────────────
    if file_type == "store_master":
        result = parse_storemaster_file(tmp_path)
        stores = result["stores"]
        if not stores:
            return {"filename": filename, "status": "⚠️ Empty", "type": "Store Master",
                    "message": "No stores found", "rows": 0}

        with get_db() as conn:
            batch_id = log_upload(conn, filename, "store_master", None,
                                  len(stores), len(stores), 0, None, None, None)
            for s in stores:
                upsert_store_master(
                    conn, s["gold_code"], s["pos_code"], s["store_name"],
                    s["segment_id"], s["segment_name"], s["company_code"],
                    s["legal_entity"], s["channel"], s["address_state"],
                    s["address_township"], s["latitude"], s["longitude"],
                    s["store_size"], s["open_date"], s["closed_date"], s["sector_id"],
                )

        sector_info = ", ".join(f"{k}: {v}" for k, v in result["sector_counts"].items())
        return {"filename": filename, "status": "✅ Imported", "type": "Store Master",
                "message": f"{len(stores)} stores ({sector_info})",
                "rows": len(stores), "data": pd.DataFrame(stores[:20])}

    # ─── Daily Sales ─────────────────────────────────────────────────────
    if file_type == "daily_sales":
        result = parse_daily_sales_file(tmp_path)
        records = result["records"]
        if not records:
            return {"filename": filename, "status": "⚠️ Empty", "type": "Daily Sales",
                    "message": "No sales records found", "rows": 0}

        with get_db() as conn:
            batch_id = log_upload(conn, filename, "daily_sales", None,
                                  len(records), len(records), len(result["errors"]),
                                  result["date_range"][0], result["date_range"][1],
                                  result["errors"][:50] if result["errors"] else None)
            for r in records:
                upsert_daily_sale(
                    conn, r["sales_site_name"], r.get("sector_id"), r["date"],
                    r["brand"], r["sales_amt"], r["margin"],
                    source="excel", upload_batch_id=batch_id,
                )

        dr = result["date_range"]
        return {"filename": filename, "status": "✅ Imported", "type": "Daily Sales",
                "message": f"{len(records)} records, {len(result['sites_found'])} sites, "
                           f"{len(result['brands_found'])} brands ({dr[0]} → {dr[1]})",
                "rows": len(records), "errors": result["errors"],
                "data": pd.DataFrame(records[:20])}

    # ─── Hourly Sales ────────────────────────────────────────────────────
    if file_type == "hourly_sales":
        result = parse_hourly_sales_file(tmp_path)
        records = result["records"]
        if not records:
            return {"filename": filename, "status": "⚠️ Empty", "type": "Hourly Sales",
                    "message": "No hourly sales records found", "rows": 0}

        with get_db() as conn:
            batch_id = log_upload(conn, filename, "hourly_sales", None,
                                  len(records), len(records), len(result["errors"]),
                                  result["date_range"][0], result["date_range"][1],
                                  result["errors"][:50] if result["errors"] else None)
            for r in records:
                upsert_hourly_sale(
                    conn, r["sales_site_name"], r.get("sector_id"), r["date"],
                    r["hour"], r["brand"], r["sales_amt"], r["trans_cnt"],
                    source="excel", upload_batch_id=batch_id,
                )

        dr = result["date_range"]
        return {"filename": filename, "status": "✅ Imported", "type": "Hourly Sales",
                "message": f"{len(records)} records, {len(result['sites_found'])} sites, "
                           f"hours 0-23 ({dr[0]} → {dr[1]})",
                "rows": len(records), "errors": result["errors"],
                "data": pd.DataFrame(records[:20])}

    if file_type == "fuel_price":
        result = parse_fuel_price_file(tmp_path)
        purchases = result["purchases"]
        if not purchases:
            return {"filename": filename, "status": "⚠️ Empty", "type": "Fuel Price",
                    "message": "No fuel price records found", "rows": 0}

        # Import
        with get_db() as conn:
            batch_id = log_upload(conn, filename, "fuel_price", None,
                                  len(purchases), len(purchases), 0, None, None, None)
            for p in purchases:
                insert_fuel_purchase(
                    conn, p["sector_id"], p["date"], p["region"],
                    p["supplier"], p["fuel_type"],
                    p["quantity_liters"], p["price_per_liter"],
                    source="excel", upload_batch_id=batch_id,
                )

        sectors_found = set(p["sector_id"] for p in purchases)
        return {"filename": filename, "status": "✅ Imported", "type": "Fuel Price",
                "message": f"{len(purchases)} records across {', '.join(sectors_found)}",
                "rows": len(purchases), "data": pd.DataFrame(purchases)}

    else:
        # Blackout file
        sector_name = SECTORS.get(sector_id, {}).get("name", sector_id)
        result = parse_blackout_file(tmp_path, sector_id)
        dates = result["dates_found"]
        gens = result["generators"]
        daily = result["daily_data"]
        errors = result["errors"]

        if not daily:
            return {"filename": filename, "status": "⚠️ Empty", "type": f"Blackout {sector_id}",
                    "message": "No data rows found", "rows": 0}

        # Import
        with get_db() as conn:
            batch_id = log_upload(
                conn, filename, file_type, sector_id,
                len(daily), len(daily) - len(errors), len(errors),
                min(dates) if dates else None, max(dates) if dates else None,
                errors[:50] if errors else None,
            )
            gen_id_map = {}
            for gen in gens:
                upsert_site(conn, gen["site_id"], gen["site_name"], sector_id, gen["site_type"],
                            cost_center_code=gen.get("cost_center_code"))
                gen_id = upsert_generator(
                    conn, gen["site_id"], gen["model_name"],
                    gen["model_name_raw"], gen["power_kva"],
                    gen["consumption_per_hour"], gen["fuel_type"], gen["supplier"],
                )
                gen_id_map[(gen["site_id"], gen["model_name_raw"])] = gen_id
                conn.execute("INSERT OR IGNORE INTO generator_name_map (raw_name, canonical_name, auto_mapped) VALUES (?, ?, 1)",
                             (gen["model_name_raw"], gen["model_name"]))

            sites_dates = set()
            for row in daily:
                gen_key = (row["site_id"], row["model_name_raw"])
                gen_id = gen_id_map.get(gen_key)
                if gen_id:
                    upsert_daily_operation(
                        conn, gen_id, row["site_id"], row["date"],
                        row["gen_run_hr"], row["daily_used_liters"],
                        row["spare_tank_balance"], row["blackout_hr"],
                        source="excel", upload_batch_id=batch_id,
                    )
                    sites_dates.add((row["site_id"], row["date"]))
            for sid, dt in sites_dates:
                refresh_site_summary(conn, sid, dt)

        date_range = f"{min(dates)} → {max(dates)}" if dates else "?"
        err_msg = f" ({len(errors)} rejected)" if errors else ""
        return {
            "filename": filename, "status": "✅ Imported",
            "type": f"Blackout {sector_id} ({sector_name})",
            "message": f"{len(daily)} rows, {len(gens)} generators, {len(dates)} days ({date_range}){err_msg}",
            "rows": len(daily),
            "errors": errors,
            "data": pd.DataFrame(daily[:20]) if daily else pd.DataFrame(),
        }


def _process_combo_sales(tmp_path, filename):
    """
    Process a combo sales file that has multiple sheets:
    - 'daily sales' → daily_sales table
    - 'hourly sales' → hourly_sales table
    - 'STORE MASTER' → store_master table
    - 'mapping' → resolves GOLD_CODE → BCP site_id directly into daily_sales/hourly_sales
    """
    import openpyxl
    wb = openpyxl.load_workbook(str(tmp_path), read_only=True)
    sheets = {s.lower().strip(): s for s in wb.sheetnames}
    wb.close()

    sub_results = []
    total_rows = 0

    # ─── 0. Build CostCenter → BCP site_id lookup (direct match) ───
    # sales.CostCenter = sites.cost_center_code (from blackout file)
    cc_to_site = {}  # {cost_center_str: site_id}
    with get_db() as conn:
        for r in conn.execute("SELECT site_id, cost_center_code FROM sites WHERE cost_center_code IS NOT NULL").fetchall():
            cc_to_site[str(r["cost_center_code"])] = r["site_id"]

    # ─── 1. Daily Sales ───────────────────────────────────────────────
    daily_sheet = None
    for key, name in sheets.items():
        if "daily" in key and "sales" in key:
            daily_sheet = name
            break
    if daily_sheet:
        result = parse_daily_sales_file(tmp_path, sheet_name=daily_sheet)
        records = result["records"]
        if records:
            # Read CostCenter column from raw Excel to resolve site_id
            raw_df = pd.read_excel(str(tmp_path), sheet_name=daily_sheet)
            cc_col = None
            for c in raw_df.columns:
                if "costcenter" in str(c).lower().replace(" ", "").replace("_", ""):
                    cc_col = c
                    break
            # Build row-level CostCenter lookup (by index)
            row_cc = {}
            if cc_col is not None:
                for idx, val in raw_df[cc_col].items():
                    if pd.notna(val):
                        row_cc[idx] = str(int(val))

            mapped_count = 0
            with get_db() as conn:
                batch_id = log_upload(conn, filename, "daily_sales", None,
                                      len(records), len(records), len(result["errors"]),
                                      result["date_range"][0], result["date_range"][1], None)
                for idx, r in enumerate(records):
                    # Resolve site_id: CostCenter from sales row → sites.cost_center_code
                    cc_val = row_cc.get(idx)
                    site_id = cc_to_site.get(cc_val) if cc_val else None
                    if site_id:
                        mapped_count += 1
                    upsert_daily_sale(conn, r["sales_site_name"], r.get("sector_id"),
                                     r["date"], r["brand"], r["sales_amt"], r["margin"],
                                     source="excel", upload_batch_id=batch_id,
                                     site_id=site_id)
            dr = result["date_range"]
            sub_results.append(f"Daily Sales: {len(records):,} rows, {len(result['sites_found'])} sites, "
                               f"{mapped_count} mapped to BCP ({dr[0]} → {dr[1]})")
            total_rows += len(records)

    # ─── 2. Hourly Sales ──────────────────────────────────────────────
    hourly_sheet = None
    for key, name in sheets.items():
        if "hourly" in key and "sales" in key:
            hourly_sheet = name
            break
    if hourly_sheet:
        result = parse_hourly_sales_file(tmp_path, sheet_name=hourly_sheet)
        records = result["records"]
        if records:
            # Read CostCenter column from raw Excel
            raw_hdf = pd.read_excel(str(tmp_path), sheet_name=hourly_sheet)
            hcc_col = None
            for c in raw_hdf.columns:
                if "costcenter" in str(c).lower().replace(" ", "").replace("_", ""):
                    hcc_col = c
                    break
            hrow_cc = {}
            if hcc_col is not None:
                for idx, val in raw_hdf[hcc_col].items():
                    if pd.notna(val):
                        hrow_cc[idx] = str(int(val))

            mapped_h = 0
            with get_db() as conn:
                batch_id = log_upload(conn, filename, "hourly_sales", None,
                                      len(records), len(records), len(result["errors"]),
                                      result["date_range"][0], result["date_range"][1], None)
                for idx, r in enumerate(records):
                    cc_val = hrow_cc.get(idx)
                    site_id = cc_to_site.get(cc_val) if cc_val else None
                    if site_id:
                        mapped_h += 1
                    upsert_hourly_sale(conn, r["sales_site_name"], r.get("sector_id"),
                                      r["date"], r["hour"], r["brand"], r["sales_amt"],
                                      r["trans_cnt"], source="excel", upload_batch_id=batch_id,
                                      site_id=site_id)
            dr = result["date_range"]
            sub_results.append(f"Hourly Sales: {len(records):,} rows, {mapped_h} mapped ({dr[0]} → {dr[1]})")
            total_rows += len(records)

    # ─── 3. Store Master ──────────────────────────────────────────────
    master_sheet = None
    for key, name in sheets.items():
        if "store master" in key or "storemaster" in key:
            master_sheet = name
            break
    if master_sheet:
        try:
            sm_df = pd.read_excel(str(tmp_path), sheet_name=master_sheet)
            col_map = {}
            for col in sm_df.columns:
                cl = str(col).strip().lower().replace(" ", "").replace("_", "")
                if "goldcode" in cl:
                    col_map["gold"] = col
                elif "poscode" in cl:
                    col_map["pos"] = col
                elif cl == "costcentername" or cl == "costcenter name":
                    col_map["name"] = col
                    col_map["cc_name"] = col
                elif cl == "costcenter" or (cl.startswith("costcenter") and "name" not in cl and "description" not in cl and "category" not in cl):
                    col_map["cc_code"] = col
                elif "segmentname" in cl:
                    col_map["segment_name"] = col
                elif cl == "segment":
                    col_map["segment_id"] = col
                elif "companycode" in cl:
                    col_map["company"] = col
                elif "legalentity" in cl:
                    col_map["legal"] = col
                elif cl == "channel":
                    col_map["channel"] = col
                elif "addressstate" in cl:
                    col_map["state"] = col
                elif "addresstownship" in cl:
                    col_map["township"] = col
                elif cl == "latitude":
                    col_map["lat"] = col
                elif cl == "longitude":
                    col_map["lng"] = col
                elif "storesize" in cl:
                    col_map["size"] = col
                elif "opendate" in cl:
                    col_map["open"] = col
                elif "closeddate" in cl:
                    col_map["closed"] = col
                elif cl == "sector":
                    col_map["sector"] = col

            from config.settings import SEGMENT_SECTOR_MAP
            count = 0
            with get_db() as conn:
                batch_id = log_upload(conn, filename, "store_master", None,
                                      len(sm_df), len(sm_df), 0, None, None, None)
                for _, row in sm_df.iterrows():
                    gold = str(row.get(col_map.get("gold", ""), "")).strip()
                    if not gold or gold == "nan":
                        continue
                    # Convert float GOLD_Code to int string
                    try:
                        gold = str(int(float(gold)))
                    except (ValueError, OverflowError):
                        pass
                    pos = str(row.get(col_map.get("pos", ""), "")).strip()
                    name = str(row.get(col_map.get("name", ""), "")).strip()
                    seg_name = str(row.get(col_map.get("segment_name", ""), "")).strip()
                    seg_id = row.get(col_map.get("segment_id", ""))
                    company = str(row.get(col_map.get("company", ""), "")).strip()
                    legal = str(row.get(col_map.get("legal", ""), "")).strip()
                    channel = str(row.get(col_map.get("channel", ""), "")).strip()
                    state = str(row.get(col_map.get("state", ""), "")).strip()
                    township = str(row.get(col_map.get("township", ""), "")).strip()
                    lat = row.get(col_map.get("lat")) if col_map.get("lat") else None
                    lng = row.get(col_map.get("lng")) if col_map.get("lng") else None
                    size = str(row.get(col_map.get("size", ""), "")).strip()
                    open_d = str(row.get(col_map.get("open", ""), "")).strip()
                    closed_d = str(row.get(col_map.get("closed", ""), "")).strip()
                    sector = SEGMENT_SECTOR_MAP.get(seg_name, str(row.get(col_map.get("sector", ""), "")).strip())
                    if not sector or sector == "nan":
                        sector = None

                    cc_raw = row.get(col_map.get("cc_code", ""))
                    cc_code = str(int(cc_raw)) if pd.notna(cc_raw) and cc_raw else None
                    cc_name_val = str(row.get(col_map.get("cc_name", ""), "")).strip()
                    cc_name_val = cc_name_val if cc_name_val and cc_name_val != "nan" else None

                    upsert_store_master(conn, gold, pos if pos != "nan" else None, name,
                                       seg_id, seg_name, company, legal, channel, state,
                                       township, lat, lng, size if size != "nan" else None,
                                       open_d if open_d != "nan" else None,
                                       closed_d if closed_d != "nan" else None, sector,
                                       cost_center_code=cc_code, cost_center_name=cc_name_val)
                    count += 1
            sub_results.append(f"Store Master: {count} stores")
            total_rows += count
        except Exception as e:
            sub_results.append(f"Store Master: error — {e}")

    if not sub_results:
        return {"filename": filename, "status": "⚠️ Empty", "type": "Combo Sales",
                "message": "No data found in any sheet", "rows": 0}

    return {
        "filename": filename, "status": "✅ Imported", "type": "Combo Sales (multi-sheet)",
        "message": " | ".join(sub_results),
        "rows": total_rows,
    }


# ═══════════════════════════════════════════════════════════════════════════
# TAB 1: Upload Files — Two sections: Weekly + Reference
# ═══════════════════════════════════════════════════════════════════════════
if selected == "📁 Upload Files":

    upload_section = ui.tabs(
        options=["📅 Daily Data", "📚 Reference Data"],
        default_value="📅 Daily Data",
        key="upload_section_tabs",
    )

    # ─── Helper: process one uploaded file and show result ─────────────
    # Map expected_type to what detect_file_type() returns
    SLOT_TYPE_MAP = {
        "blackout_cmhl": ("blackout_cmhl",),
        "blackout_cp": ("blackout_cp",),
        "blackout_cfc": ("blackout_cfc",),
        "fuel_price": ("fuel_price",),
        "daily_sales": ("daily_sales", "combo_sales"),
        "hourly_sales": ("hourly_sales",),
        "combo_sales": ("combo_sales",),
        "store_master": ("store_master",),
        "site_mapping": ("site_mapping",),
        "diesel_expense_ly": ("diesel_expense_ly",),
    }

    # Friendly names for error messages
    SLOT_FRIENDLY = {
        "blackout_cmhl": "CMHL Blackout (sheet 'CMHL')",
        "blackout_cp": "CP Blackout (sheet 'CP')",
        "blackout_cfc": "CFC Blackout (sheet 'CFC')",
        "fuel_price": "Daily Fuel Price (sheets CMHL/CP/CFC/PG)",
        "daily_sales": "Daily Sales (sheet 'daily sales')",
        "hourly_sales": "Hourly Sales (sheet 'hourly sales')",
        "combo_sales": "Combo Sales (sheets 'daily sales' + 'hourly sales')",
        "store_master": "Store Master",
        "site_mapping": "Site Mapping (columns 'Manual Data' + 'SAP Cost Center')",
        "diesel_expense_ly": "LY Diesel Expense (column 'Daily Average Diesel Exp')",
    }

    def _get_last_sync(file_type):
        """Get last upload date and row count for a file type."""
        with get_db() as conn:
            # For combo_sales, check any of the sub-types it logs as
            if file_type == "combo_sales":
                row = conn.execute(
                    "SELECT filename, rows_accepted, uploaded_at FROM upload_history "
                    "WHERE file_type IN ('daily_sales', 'hourly_sales', 'store_master', 'combo_sales') "
                    "ORDER BY uploaded_at DESC LIMIT 1"
                ).fetchone()
            else:
                row = conn.execute(
                    "SELECT filename, rows_accepted, uploaded_at FROM upload_history "
                    "WHERE file_type = ? ORDER BY uploaded_at DESC LIMIT 1", (file_type,)
                ).fetchone()
        if row:
            return row["filename"], row["rows_accepted"], row["uploaded_at"][:16] if row["uploaded_at"] else ""
        return None, 0, None

    def _get_table_preview(query, limit=10):
        """Get a preview dataframe from DB."""
        with get_db() as conn:
            return pd.read_sql_query(query + f" LIMIT {limit}", conn)

    # Table queries for preview after each upload type
    PREVIEW_QUERIES = {
        "blackout_cmhl": "SELECT do.date, do.site_id, g.model_name, do.gen_run_hr, do.daily_used_liters, do.spare_tank_balance FROM daily_operations do JOIN generators g ON do.generator_id = g.generator_id JOIN sites s ON do.site_id = s.site_id WHERE s.sector_id = 'CMHL' ORDER BY do.date DESC, do.site_id",
        "blackout_cp": "SELECT do.date, do.site_id, g.model_name, do.gen_run_hr, do.daily_used_liters, do.spare_tank_balance FROM daily_operations do JOIN generators g ON do.generator_id = g.generator_id JOIN sites s ON do.site_id = s.site_id WHERE s.sector_id = 'CP' ORDER BY do.date DESC, do.site_id",
        "blackout_cfc": "SELECT do.date, do.site_id, g.model_name, do.gen_run_hr, do.daily_used_liters, do.spare_tank_balance FROM daily_operations do JOIN generators g ON do.generator_id = g.generator_id JOIN sites s ON do.site_id = s.site_id WHERE s.sector_id = 'CFC' ORDER BY do.date DESC, do.site_id",
        "fuel_price": "SELECT date, sector_id, region, supplier, fuel_type, quantity_liters, price_per_liter FROM fuel_purchases ORDER BY date DESC",
        "combo_sales": None,  # Special handling — shows multiple tables
        "diesel_expense_ly": "SELECT cost_center_code, sector_id, cost_center_name, daily_avg_expense_mmk, pct_on_sales FROM diesel_expense_ly ORDER BY daily_avg_expense_mmk DESC",
        "store_master": "SELECT gold_code, store_name, segment_name, sector_id, channel, address_state FROM store_master ORDER BY sector_id",
        "site_mapping": "SELECT sales_site_name, site_id, sector_id, match_method FROM site_sales_map ORDER BY sector_id",
    }

    def _upload_slot(label, description, expected_type, key_prefix):
        """Render a full-width upload row with last sync info, upload, and data preview."""
        # Header with last sync
        _, last_rows, last_sync = _get_last_sync(expected_type)

        col_h1, col_h2 = st.columns([7, 3])
        with col_h1:
            st.markdown(f"#### {label}")
            st.caption(description)
        with col_h2:
            if last_sync:
                st.markdown(f"""
                <div style="background:#f0fdf4;border:1px solid #bbf7d0;border-radius:8px;padding:8px 12px;text-align:center;margin-top:8px">
                    <div style="font-size:11px;color:#16a34a;font-weight:600">Last Sync</div>
                    <div style="font-size:13px;font-weight:700">{last_sync}</div>
                    <div style="font-size:11px;color:#64748b">{last_rows:,} rows</div>
                </div>""", unsafe_allow_html=True)
            else:
                st.markdown(f"""
                <div style="background:#fef2f2;border:1px solid #fecaca;border-radius:8px;padding:8px 12px;text-align:center;margin-top:8px">
                    <div style="font-size:11px;color:#dc2626;font-weight:600">Not Uploaded</div>
                    <div style="font-size:13px;color:#64748b">No data yet</div>
                </div>""", unsafe_allow_html=True)

        # Clear queries per slot — clears only this sector's data
        CLEAR_QUERIES = {
            "blackout_cmhl": [
                "DELETE FROM daily_operations WHERE site_id IN (SELECT site_id FROM sites WHERE sector_id='CMHL')",
                "DELETE FROM daily_site_summary WHERE site_id IN (SELECT site_id FROM sites WHERE sector_id='CMHL')",
            ],
            "blackout_cp": [
                "DELETE FROM daily_operations WHERE site_id IN (SELECT site_id FROM sites WHERE sector_id='CP')",
                "DELETE FROM daily_site_summary WHERE site_id IN (SELECT site_id FROM sites WHERE sector_id='CP')",
            ],
            "blackout_cfc": [
                "DELETE FROM daily_operations WHERE site_id IN (SELECT site_id FROM sites WHERE sector_id='CFC')",
                "DELETE FROM daily_site_summary WHERE site_id IN (SELECT site_id FROM sites WHERE sector_id='CFC')",
            ],
            "fuel_price": [
                "DELETE FROM fuel_purchases",
            ],
            "combo_sales": [
                "DELETE FROM daily_sales",
                "DELETE FROM hourly_sales",
            ],
        }

        # Upload + action buttons
        uploaded = st.file_uploader(
            f"Upload {label}", type=["xlsx"], key=f"up_{key_prefix}",
            label_visibility="collapsed",
        )

        if uploaded:
            do_replace = st.button(
                f"🔄 Import {uploaded.name}", key=f"btn_replace_{key_prefix}",
                type="primary", use_container_width=True,
            )

            if do_replace:
                with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
                    tmp.write(uploaded.read())
                    tmp_path = tmp.name

                # Validate file type matches this slot
                detected_type, detected_sector = detect_file_type(tmp_path)
                allowed = SLOT_TYPE_MAP.get(expected_type, (expected_type,))

                if detected_type not in allowed:
                    os.unlink(tmp_path)
                    detected_name = SLOT_FRIENDLY.get(detected_type, detected_type or "Unknown")
                    expected_name = SLOT_FRIENDLY.get(expected_type, expected_type)
                    st.error(
                        f"**Wrong file!** This slot expects **{expected_name}** "
                        f"but your file was detected as **{detected_name}**. "
                        f"Please upload the correct file."
                    )
                else:
                    # Clear old data for this sector before importing
                    clear_sqls = CLEAR_QUERIES.get(expected_type, [])
                    if clear_sqls:
                        with get_db() as conn:
                            for sql in clear_sqls:
                                conn.execute(sql)
                            conn.execute("DELETE FROM alerts")
                            conn.execute("DELETE FROM ai_insights_cache")

                    with st.spinner(f"Processing {uploaded.name}..."):
                        result = process_file(tmp_path, uploaded.name)
                    os.unlink(tmp_path)

                    if "✅" in result["status"]:
                        if expected_type in ("blackout_cmhl", "blackout_cp", "blackout_cfc", "fuel_price"):
                            try:
                                run_all_checks()
                            except Exception:
                                pass
                        # Backfill site_id in sales tables using CostCenter→cost_center_code
                        try:
                            with get_db() as conn:
                                # Build cc → site_id from current sites
                                cc_map = {r[0]: r[1] for r in conn.execute(
                                    "SELECT cost_center_code, site_id FROM sites WHERE cost_center_code IS NOT NULL"
                                ).fetchall()}
                                if cc_map:
                                    # Also build gold_code → cc from store_master
                                    gold_cc = {r[0]: r[1] for r in conn.execute(
                                        "SELECT gold_code, cost_center_code FROM store_master WHERE cost_center_code IS NOT NULL"
                                    ).fetchall()}
                                    updated = 0
                                    for gold, cc in gold_cc.items():
                                        site_id = cc_map.get(cc)
                                        if site_id:
                                            c1 = conn.execute(
                                                "UPDATE daily_sales SET site_id = ? WHERE sales_site_name = ? AND site_id IS NULL",
                                                (site_id, gold)).rowcount
                                            c2 = conn.execute(
                                                "UPDATE hourly_sales SET site_id = ? WHERE sales_site_name = ? AND site_id IS NULL",
                                                (site_id, gold)).rowcount
                                            updated += c1 + c2
                                    # Also direct match: sales CostCenter stored in sales_site_name might not apply,
                                    # but catch any rows where sales_site_name IS a cost_center_code
                                    for cc, site_id in cc_map.items():
                                        conn.execute(
                                            "UPDATE daily_sales SET site_id = ? WHERE sales_site_name = ? AND site_id IS NULL",
                                            (site_id, cc))
                                        conn.execute(
                                            "UPDATE hourly_sales SET site_id = ? WHERE sales_site_name = ? AND site_id IS NULL",
                                            (site_id, cc))
                        except Exception:
                            pass
                        st.session_state[f"import_msg_{key_prefix}"] = ("success", result["message"])
                        st.rerun()
                    elif "❌" in result["status"]:
                        st.error(f"❌ {result['message']}")
                    else:
                        st.warning(result["message"])

                    if result.get("errors"):
                        with st.expander(f"🧹 {len(result['errors'])} cleaning actions"):
                            for e in result["errors"][:10]:
                                st.caption(f"- {e}")

        # Show success message from previous import (survives rerun)
        msg_key = f"import_msg_{key_prefix}"
        if msg_key in st.session_state:
            msg_type, msg_text = st.session_state.pop(msg_key)
            if msg_type == "success":
                st.success(f"✅ {msg_text}")

        # Data preview table (always show if data exists)
        if expected_type == "combo_sales":
            # Show all tables from combo import
            combo_tables = [
                ("📊 Daily Sales", "SELECT ds.date, ds.sales_site_name as site_code, sm.cost_center_code, sm.store_name, ds.sector_id, ds.brand, ds.sales_amt, ds.margin FROM daily_sales ds LEFT JOIN store_master sm ON ds.sales_site_name = sm.gold_code ORDER BY ds.date DESC"),
                ("🕐 Hourly Sales", "SELECT hs.date, hs.hour, hs.sales_site_name as site_code, sm.cost_center_code, hs.sector_id, hs.brand, hs.sales_amt, hs.trans_cnt FROM hourly_sales hs LEFT JOIN store_master sm ON hs.sales_site_name = sm.gold_code ORDER BY hs.date DESC, hs.hour"),
                ("🏪 Store Master", "SELECT gold_code, store_name, cost_center_code, cost_center_name, segment_name, sector_id FROM store_master ORDER BY sector_id"),
                ("🗺️ Sales-BCP Mapping", "SELECT sm.sales_site_name as sales_code, st.cost_center_code, sm.site_id as bcp_site, sm.sector_id, sm.match_method FROM site_sales_map sm LEFT JOIN store_master st ON sm.sales_site_name = st.gold_code ORDER BY sm.sector_id"),
            ]
            for tbl_label, tbl_query in combo_tables:
                tbl_df = _get_table_preview(tbl_query, limit=15)
                if not tbl_df.empty:
                    with st.expander(f"{tbl_label} ({len(tbl_df)}+ rows)", expanded=False):
                        st.dataframe(tbl_df, use_container_width=True, hide_index=True, height=250)
        else:
            preview_query = PREVIEW_QUERIES.get(expected_type)
            if preview_query:
                preview_df = _get_table_preview(preview_query)
                if not preview_df.empty:
                    with st.expander(f"📋 Current data ({len(preview_df)}+ rows)", expanded=False):
                        st.dataframe(preview_df, use_container_width=True, hide_index=True, height=250)

        st.markdown("---")

    # ─────────────────────────────────────────────────────────────────────
    # SECTION 1: Daily Data (uploaded daily)
    # ─────────────────────────────────────────────────────────────────────
    if upload_section == "📅 Daily Data":
        st.markdown("### 📅 Daily Data — Upload Every Day")
        st.markdown("These 5 files are updated daily. Upload the latest version to refresh dashboards.")
        st.info("Data is **upserted** — same date+site overwrites, new dates are added. No data loss.")

        st.markdown("---")

        # Blackout files — one per sector, each in its own row
        st.markdown("## 1️⃣ Blackout & Generator Files")
        st.caption("Upload daily. Generator run hours, diesel usage, tank balance, and blackout hours per store.")

        _upload_slot(
            "🏢 CMHL — City Mart Holdings",
            "31 stores, 35 generators. Sheet: 'CMHL'. Columns: Store code, Cost Center, Machine Power, per-date: Gen Run Hr, Daily Used (L), Tank Balance.",
            "blackout_cmhl", "bl_cmhl",
        )

        _upload_slot(
            "🏬 CP — City Properties",
            "25 stores, 46 generators. Sheet: 'CP'. Includes Blackout Hr. Sites: City Mart, City Mall, Marketplace.",
            "blackout_cp", "bl_cp",
        )

        _upload_slot(
            "🍞 CFC — City Food Concept",
            "2 stores (Seasons Bakery Factory, Mandalay). Sheet: 'CFC'. 4 generators at SBFTY, 1 at BMDY.",
            "blackout_cfc", "bl_cfc",
        )

        # Fuel Price
        st.markdown("## 2️⃣ Daily Fuel Price")
        _upload_slot(
            "⛽ Daily Fuel Price",
            "Diesel purchase prices from Denko and Moon Sun suppliers. "
            "Excel has 4 sheets: CMHL, CP, CFC, PG. "
            "Columns: Date, Region (YGN/MDY), Supplier, Fuel Type (PD/HSD), Quantity (L), Price/L (MMK).",
            "fuel_price", "fuel",
        )

        # Sales files
        st.markdown("## 3️⃣ Sales Data")
        st.caption("Upload daily. Compare diesel cost vs revenue. You can upload one combo file OR individual files.")

        _upload_slot(
            "📦 Sales File (All-in-One)",
            "One Excel file with sheets: 'daily sales' (32K rows), 'hourly sales' (147K rows), "
            "'STORE MASTER' (439 stores), 'mapping' (site mappings). "
            "Example: CMHL_DAILY_SALES.xlsx — one upload imports everything.",
            "combo_sales", "combo_sales",
        )

        # Clear all button
        st.markdown("---")
        if st.button("🗑️ Clear ALL Daily Data & Start Fresh", key="btn_clear_all",
                      use_container_width=True, type="secondary"):
            with get_db() as conn:
                conn.execute("DELETE FROM daily_operations")
                conn.execute("DELETE FROM daily_site_summary")
                conn.execute("DELETE FROM fuel_purchases")
                conn.execute("DELETE FROM daily_sales")
                conn.execute("DELETE FROM hourly_sales")
                conn.execute("DELETE FROM alerts")
                conn.execute("DELETE FROM ai_insights_cache")
            st.success("All daily data cleared. Upload new files above.")
            st.rerun()

    # ─────────────────────────────────────────────────────────────────────
    # SECTION 2: Reference Data (upload once, rarely changes)
    # ─────────────────────────────────────────────────────────────────────
    elif upload_section == "📚 Reference Data":
        st.markdown("### 📚 Reference Data — Upload Once")
        st.markdown("This file rarely changes. Upload once — stays across daily refreshes.")

        st.markdown("---")
        _upload_slot(
            "💰 Last Year Diesel Expense",
            "Historical baseline — last 12-month average daily diesel expense per store. "
            "3 sheets: CMHL, CP, CFC. Joined to blackout data via Cost Center Code. "
            "Used to compare current vs LY spending.",
            "diesel_expense_ly", "ref_ly",
        )

# ═══════════════════════════════════════════════════════════════════════════
# TAB: Store Summary (Excel-like pivot view — dates as columns)
# ═══════════════════════════════════════════════════════════════════════════
elif selected == "📊 Store Summary":
    st.markdown("### Store Summary — Dates as Columns")
    st.caption("Same layout as your Excel file. Aggregated per store per date.")

    with get_db() as conn:
        # Check if data exists
        check = conn.execute("SELECT COUNT(*) FROM daily_site_summary").fetchone()[0]
        if check == 0:
            st.info("No data yet. Upload blackout files first.")
            st.stop()

        # Load site summary with sector
        summary_df = pd.read_sql_query("""
            SELECT dss.date, dss.site_id, s.sector_id, s.site_name,
                   dss.total_gen_run_hr, dss.total_daily_used,
                   dss.spare_tank_balance, dss.blackout_hr,
                   dss.days_of_buffer, dss.num_generators_active
            FROM daily_site_summary dss
            JOIN sites s ON dss.site_id = s.site_id
            ORDER BY s.sector_id, dss.site_id, dss.date
        """, conn)

    if summary_df.empty:
        st.info("No summary data available.")
        st.stop()

    # Sector filter
    sectors = sorted(summary_df["sector_id"].unique().tolist())
    sector_sel = st.selectbox("Sector", ["All"] + sectors, key="ss_sector")
    if sector_sel != "All":
        summary_df = summary_df[summary_df["sector_id"] == sector_sel]

    # Short date for column headers
    summary_df["short_date"] = pd.to_datetime(summary_df["date"]).dt.strftime("%d-%b")

    # Build site label: "SECTOR-CODE | Name"
    summary_df["store"] = summary_df["site_id"]

    # Metrics to pivot
    metrics = {
        "⚙️ Generator Run Hours": {"col": "total_gen_run_hr", "fmt": "{:.1f}", "desc": "Total generator run hours per store per day"},
        "⛽ Daily Diesel Used (L)": {"col": "total_daily_used", "fmt": "{:,.0f}", "desc": "Total liters consumed per store per day"},
        "🛢️ Spare Tank Balance (L)": {"col": "spare_tank_balance", "fmt": "{:,.0f}", "desc": "Remaining diesel in tank (site-level, not per generator)"},
        "⚡ Blackout Hours": {"col": "blackout_hr", "fmt": "{:.1f}", "desc": "Hours of power outage (only CP and CFC)"},
        "📅 Days of Buffer": {"col": "days_of_buffer", "fmt": "{:.1f}", "desc": "Tank balance ÷ daily usage = days of diesel left"},
        "🔧 Active Generators": {"col": "num_generators_active", "fmt": "{:.0f}", "desc": "Generators with run hours > 0"},
    }

    metric_sel = ui.tabs(
        options=list(metrics.keys()),
        default_value="⛽ Daily Diesel Used (L)",
        key="ss_metric",
    )

    m = metrics[metric_sel]
    st.caption(m["desc"])

    # Pivot: stores as rows, dates as columns
    pivot = summary_df.pivot_table(
        index="store", columns="short_date", values=m["col"], aggfunc="sum"
    )

    # Order columns chronologically
    date_order = summary_df.sort_values("date")["short_date"].unique().tolist()
    pivot = pivot.reindex(columns=date_order)

    # Add row total/average
    if m["col"] in ("total_gen_run_hr", "total_daily_used"):
        pivot["TOTAL"] = pivot.sum(axis=1)
        pivot["AVG"] = pivot.drop(columns=["TOTAL"]).mean(axis=1)
    elif m["col"] in ("spare_tank_balance", "days_of_buffer"):
        pivot["LATEST"] = pivot.iloc[:, -1] if len(pivot.columns) > 0 else None
    elif m["col"] == "num_generators_active":
        pivot["MAX"] = pivot.max(axis=1)

    # Format values
    styled = pivot.copy()
    for col in styled.columns:
        styled[col] = styled[col].apply(
            lambda x: m["fmt"].format(x) if pd.notna(x) and x != 0 else "—"
        )

    # Color coding for buffer days
    if m["col"] == "days_of_buffer":
        def color_buffer(val):
            try:
                v = float(val.replace(",", ""))
                if v < 3:
                    return "background-color: #fecaca"  # red
                elif v < 7:
                    return "background-color: #fef3c7"  # yellow
                else:
                    return "background-color: #d1fae5"  # green
            except (ValueError, AttributeError):
                return ""
        st.dataframe(
            styled.style.applymap(color_buffer),
            use_container_width=True, height=min(50 + len(styled) * 35, 600),
        )
    else:
        st.dataframe(styled, use_container_width=True, height=min(50 + len(styled) * 35, 600))

    # Show count
    st.caption(f"{len(pivot)} stores × {len(date_order)} dates")

    # ─── LY Diesel Expense Comparison ─────────────────────────────────
    st.markdown("---")
    st.markdown("### 💰 Current vs Last Year Diesel Expense")
    st.caption("Compares current average daily diesel cost against last 12-month historical average (from LY file).")

    with get_db() as conn:
        ly_count = conn.execute("SELECT COUNT(*) FROM diesel_expense_ly").fetchone()[0]
        if ly_count == 0:
            st.info("No LY data. Upload 'Daily Avg Diesel Expense LY.xlsx' to see comparison.")
        else:
            # Get fuel price for cost calculation
            price_row = conn.execute(
                "SELECT AVG(price_per_liter) FROM fuel_purchases WHERE price_per_liter IS NOT NULL"
            ).fetchone()
            fuel_price = price_row[0] if price_row and price_row[0] else 3400  # fallback

            # Join: sites → cost_center_code → diesel_expense_ly
            # Current avg = avg daily used liters × fuel price
            compare_df = pd.read_sql_query("""
                SELECT s.site_id, s.sector_id, s.site_name, s.cost_center_code,
                       AVG(dss.total_daily_used) as avg_daily_liters,
                       de.daily_avg_expense_mmk as ly_avg_mmk,
                       de.yearly_expense_mmk_mil as ly_yearly_mil,
                       de.pct_on_sales as ly_pct_sales
                FROM daily_site_summary dss
                JOIN sites s ON dss.site_id = s.site_id
                LEFT JOIN diesel_expense_ly de ON s.cost_center_code = de.cost_center_code
                WHERE dss.total_daily_used > 0
                GROUP BY s.site_id
                ORDER BY s.sector_id, s.site_id
            """, conn)

            if sector_sel != "All":
                compare_df = compare_df[compare_df["sector_id"] == sector_sel]

            if compare_df.empty:
                st.info("No matched data for comparison.")
            else:
                # Calculate current daily cost
                compare_df["current_avg_mmk"] = compare_df["avg_daily_liters"] * fuel_price
                compare_df["has_ly"] = compare_df["ly_avg_mmk"].notna()
                compare_df["delta_mmk"] = compare_df["current_avg_mmk"] - compare_df["ly_avg_mmk"]
                compare_df["delta_pct"] = (compare_df["delta_mmk"] / compare_df["ly_avg_mmk"] * 100)

                # KPI cards
                matched = compare_df[compare_df["has_ly"]]
                if not matched.empty:
                    avg_increase = matched["delta_pct"].mean()
                    total_current = matched["current_avg_mmk"].sum()
                    total_ly = matched["ly_avg_mmk"].sum()

                    kc1, kc2, kc3, kc4 = st.columns(4)
                    with kc1:
                        ui.metric_card(title="Stores Matched", content=str(len(matched)),
                                       description=f"of {len(compare_df)} total", key="ly_matched")
                    with kc2:
                        color = "red" if avg_increase > 0 else "green"
                        ui.metric_card(title="Avg Change",
                                       content="{:+.0f}%".format(avg_increase),
                                       description="vs LY average", key="ly_avg_chg")
                    with kc3:
                        ui.metric_card(title="Current Total/day",
                                       content="{:,.0f}".format(total_current),
                                       description="MMK/day (all matched)", key="ly_cur_total")
                    with kc4:
                        ui.metric_card(title="LY Total/day",
                                       content="{:,.0f}".format(total_ly),
                                       description="MMK/day (LY avg)", key="ly_ly_total")

                # Display table
                display = compare_df[["site_id", "sector_id", "avg_daily_liters",
                                       "current_avg_mmk", "ly_avg_mmk", "delta_mmk", "delta_pct"]].copy()
                display.columns = ["Store", "Sector", "Avg Daily (L)", "Current (MMK/day)",
                                    "LY Avg (MMK/day)", "Δ MMK/day", "Δ %"]

                # Format
                display["Avg Daily (L)"] = display["Avg Daily (L)"].apply(
                    lambda x: "{:,.0f}".format(x) if pd.notna(x) else "—")
                display["Current (MMK/day)"] = display["Current (MMK/day)"].apply(
                    lambda x: "{:,.0f}".format(x) if pd.notna(x) else "—")
                display["LY Avg (MMK/day)"] = display["LY Avg (MMK/day)"].apply(
                    lambda x: "{:,.0f}".format(x) if pd.notna(x) else "No LY data")
                display["Δ MMK/day"] = display["Δ MMK/day"].apply(
                    lambda x: "{:+,.0f}".format(x) if pd.notna(x) else "—")
                display["Δ %"] = display["Δ %"].apply(
                    lambda x: "{:+.0f}%".format(x) if pd.notna(x) else "—")

                st.dataframe(display, use_container_width=True, hide_index=True,
                             height=min(50 + len(display) * 35, 600))
                st.caption("Fuel price used: {:,.0f} MMK/L".format(fuel_price))

    # ─── Sales vs Diesel by Sector & Date ─────────────────────────────
    st.markdown("---")
    st.markdown("### 💰 Sales vs Diesel Cost — By Sector & Date")
    st.caption("Compares daily sales revenue against diesel cost for the same dates. Shows diesel as % of sales.")

    with get_db() as conn:
        sales_check = conn.execute("SELECT COUNT(*) FROM daily_sales").fetchone()[0]
        diesel_check = conn.execute("SELECT COUNT(*) FROM daily_site_summary").fetchone()[0]

        if sales_check == 0:
            st.info("No sales data. Upload sales file (combo or individual) to see comparison.")
        elif diesel_check == 0:
            st.info("No diesel data. Upload blackout files first.")
        else:
            # Get fuel price
            price_row = conn.execute(
                "SELECT AVG(price_per_liter) FROM fuel_purchases WHERE price_per_liter IS NOT NULL"
            ).fetchone()
            fuel_price = price_row[0] if price_row and price_row[0] else 3400

            # Diesel cost by sector + date
            diesel_df = pd.read_sql_query("""
                SELECT s.sector_id, dss.date,
                       SUM(dss.total_daily_used) as total_liters,
                       SUM(dss.spare_tank_balance) as total_tank,
                       COUNT(DISTINCT dss.site_id) as sites
                FROM daily_site_summary dss
                JOIN sites s ON dss.site_id = s.site_id
                WHERE dss.total_daily_used > 0
                GROUP BY s.sector_id, dss.date
                ORDER BY s.sector_id, dss.date
            """, conn)

            # Sales by sector + date
            sales_df = pd.read_sql_query("""
                SELECT sector_id, date,
                       SUM(sales_amt) as total_sales,
                       COUNT(DISTINCT sales_site_name) as sites
                FROM daily_sales
                WHERE sector_id IS NOT NULL
                GROUP BY sector_id, date
                ORDER BY sector_id, date
            """, conn)

        if sales_check > 0 and diesel_check > 0 and not diesel_df.empty and not sales_df.empty:
            # Find overlapping dates
            diesel_dates = set(diesel_df["date"])
            sales_dates = set(sales_df["date"])
            overlap_dates = sorted(diesel_dates & sales_dates)

            if not overlap_dates:
                st.warning("No overlapping dates between sales and diesel data.")
            else:
                st.success(f"Comparing {len(overlap_dates)} overlapping days: {overlap_dates[0]} → {overlap_dates[-1]}")

                # Filter to overlap
                d_filt = diesel_df[diesel_df["date"].isin(overlap_dates)].copy()
                s_filt = sales_df[sales_df["date"].isin(overlap_dates)].copy()

                d_filt["diesel_cost_mmk"] = d_filt["total_liters"] * fuel_price

                # Merge
                merged = d_filt.merge(s_filt, on=["sector_id", "date"], how="outer", suffixes=("_diesel", "_sales"))
                merged["diesel_pct"] = (merged["diesel_cost_mmk"] / merged["total_sales"] * 100).round(2)
                merged["short_date"] = pd.to_datetime(merged["date"]).dt.strftime("%d-%b")

                if sector_sel != "All":
                    merged = merged[merged["sector_id"] == sector_sel]

                # ── KPI Summary ──
                st.markdown("#### Summary (Overlapping Period)")
                sectors_in_data = sorted(merged["sector_id"].dropna().unique())

                for sector in sectors_in_data:
                    sec_data = merged[merged["sector_id"] == sector]
                    total_diesel = sec_data["diesel_cost_mmk"].sum()
                    total_sales = sec_data["total_sales"].sum()
                    avg_diesel = sec_data["diesel_cost_mmk"].mean()
                    avg_sales = sec_data["total_sales"].mean()
                    pct = (total_diesel / total_sales * 100) if total_sales > 0 else 0

                    pct_color = "#dc2626" if pct > 5 else "#d97706" if pct > 2 else "#16a34a"
                    st.markdown(f"""
                    <div style="background:#f8fafc;border:1px solid #e2e8f0;border-left:4px solid {pct_color};
                                border-radius:0 8px 8px 0;padding:12px 16px;margin:8px 0">
                        <strong>{sector}</strong> — Diesel is <strong style="color:{pct_color}">{pct:.2f}%</strong> of sales
                        <br><span style="font-size:12px;color:#64748b">
                        Total Diesel: {total_diesel:,.0f} MMK | Total Sales: {total_sales:,.0f} MMK |
                        Avg Diesel/day: {avg_diesel:,.0f} MMK | Avg Sales/day: {avg_sales:,.0f} MMK
                        </span>
                    </div>""", unsafe_allow_html=True)

                # ── Daily Pivot: Diesel Cost ──
                st.markdown("#### Daily Diesel Cost by Sector (MMK)")
                pivot_diesel = merged.pivot_table(
                    index="sector_id", columns="short_date", values="diesel_cost_mmk", aggfunc="sum"
                )
                date_order = merged.sort_values("date")["short_date"].unique().tolist()
                pivot_diesel = pivot_diesel.reindex(columns=date_order)
                pivot_diesel["TOTAL"] = pivot_diesel.sum(axis=1)
                pivot_diesel["AVG/day"] = pivot_diesel.drop(columns=["TOTAL"]).mean(axis=1)

                styled_d = pivot_diesel.copy()
                for col in styled_d.columns:
                    styled_d[col] = styled_d[col].apply(
                        lambda x: "{:,.0f}".format(x) if pd.notna(x) and x != 0 else "—")
                st.dataframe(styled_d, use_container_width=True)

                # ── Daily Pivot: Sales ──
                st.markdown("#### Daily Sales by Sector (MMK)")
                pivot_sales = merged.pivot_table(
                    index="sector_id", columns="short_date", values="total_sales", aggfunc="sum"
                )
                pivot_sales = pivot_sales.reindex(columns=date_order)
                pivot_sales["TOTAL"] = pivot_sales.sum(axis=1)
                pivot_sales["AVG/day"] = pivot_sales.drop(columns=["TOTAL"]).mean(axis=1)

                styled_s = pivot_sales.copy()
                for col in styled_s.columns:
                    styled_s[col] = styled_s[col].apply(
                        lambda x: "{:,.0f}".format(x) if pd.notna(x) and x != 0 else "—")
                st.dataframe(styled_s, use_container_width=True)

                # ── Daily Pivot: Diesel % of Sales ──
                st.markdown("#### Diesel as % of Sales (by day)")
                pivot_pct = merged.pivot_table(
                    index="sector_id", columns="short_date", values="diesel_pct", aggfunc="mean"
                )
                pivot_pct = pivot_pct.reindex(columns=date_order)
                pivot_pct["AVG"] = pivot_pct.mean(axis=1)

                styled_p = pivot_pct.copy()
                for col in styled_p.columns:
                    styled_p[col] = styled_p[col].apply(
                        lambda x: "{:.2f}%".format(x) if pd.notna(x) else "—")

                def color_pct(val):
                    try:
                        v = float(val.replace("%", ""))
                        if v > 5:
                            return "background-color: #fecaca"
                        elif v > 2:
                            return "background-color: #fef3c7"
                        else:
                            return "background-color: #d1fae5"
                    except (ValueError, AttributeError):
                        return ""

                st.dataframe(
                    styled_p.style.applymap(color_pct),
                    use_container_width=True,
                )
                st.caption("🟢 <2% | 🟡 2-5% | 🔴 >5% of sales")

# ═══════════════════════════════════════════════════════════════════════════
# TAB: Raw Data Browser
# ═══════════════════════════════════════════════════════════════════════════
elif selected == "🗄️ Raw Data":
    st.markdown("### Database Tables")
    st.caption("Browse all data stored in the system — one tab per table")

    TABLE_CONFIG = {
        "Sites": ("SELECT s.site_id, s.sector_id, s.site_type, s.site_name FROM sites s ORDER BY s.sector_id, s.site_id", "57 sites across 3 sectors (CP, CMHL, CFC)"),
        "Generators": ("SELECT g.site_id, s.sector_id, g.model_name, g.power_kva, g.consumption_per_hour, g.fuel_type, g.supplier FROM generators g JOIN sites s ON g.site_id = s.site_id ORDER BY s.sector_id, g.site_id", "86 generators — model, KVA rating, fuel consumption rate"),
        "Daily Ops": ("SELECT do.date, do.site_id, g.model_name, do.gen_run_hr, do.daily_used_liters, do.spare_tank_balance, do.blackout_hr, do.source FROM daily_operations do JOIN generators g ON do.generator_id = g.generator_id ORDER BY do.date DESC, do.site_id LIMIT 500", "Per-generator per-day: run hours, fuel used, tank balance"),
        "Site Summary": ("SELECT dss.date, dss.site_id, s.sector_id, dss.total_gen_run_hr, dss.total_daily_used, dss.spare_tank_balance, dss.days_of_buffer FROM daily_site_summary dss JOIN sites s ON dss.site_id = s.site_id ORDER BY dss.date DESC LIMIT 500", "Aggregated per-site: buffer = tank balance / daily usage"),
        "Fuel Prices": ("SELECT date, sector_id, region, supplier, fuel_type, quantity_liters, price_per_liter FROM fuel_purchases ORDER BY date DESC LIMIT 200", "Diesel prices from Denko & Moon Sun (MMK/liter)"),
        "Daily Sales": ("SELECT date, sales_site_name, sector_id, brand, sales_amt, margin FROM daily_sales ORDER BY date DESC LIMIT 500", "Daily sales per site per brand (MMK)"),
        "Hourly Sales": ("SELECT date, hour, sales_site_name, sector_id, brand, sales_amt, trans_cnt FROM hourly_sales ORDER BY date DESC, hour LIMIT 500", "Hourly sales with transaction counts"),
        "Store Master": ("SELECT gold_code, segment_name, sector_id, company_code, channel, address_state, store_size FROM store_master ORDER BY sector_id LIMIT 500", "439 stores with segment and location data"),
        "Alerts": ("SELECT alert_type, severity, site_id, sector_id, message, is_acknowledged, created_at FROM alerts ORDER BY created_at DESC LIMIT 200", "Auto-generated alerts: buffer, price, efficiency thresholds"),
        "Uploads": ("SELECT filename, file_type, sector_id, rows_parsed, rows_accepted, rows_rejected, uploaded_at FROM upload_history ORDER BY uploaded_at DESC", "Excel file upload audit trail"),
    }

    table_tab = ui.tabs(options=list(TABLE_CONFIG.keys()), default_value="Sites", key="raw_tabs")

    query, desc = TABLE_CONFIG[table_tab]

    st.caption(f"**{table_tab}:** {desc}")

    search = st.text_input("🔍 Filter rows", key=f"raw_search_{table_tab}", placeholder="Type to search...")

    with get_db() as conn:
        df = pd.read_sql_query(query, conn)

    if search:
        mask = df.astype(str).apply(lambda row: row.str.contains(search, case=False).any(), axis=1)
        df = df[mask]

    ui.metric_card(title=table_tab, content=str(len(df)), description="rows", key=f"mc_raw_{table_tab}")
    st.dataframe(df, use_container_width=True, hide_index=True, height=400)

# ═══════════════════════════════════════════════════════════════════════════
# TAB: How AI Cleans Data
# ═══════════════════════════════════════════════════════════════════════════
elif selected == "🧹 How AI Cleans Data":
    st.markdown("### How the System Handles Messy Excel Data")
    st.markdown("Your Excel files have many inconsistencies. Here's what gets fixed **automatically**:")

    st.markdown("#### 🔍 File Detection (by sheet names, NOT file names)")
    st.markdown("""
    | Sheet Names Inside File | Detected As |
    |---|---|
    | Sheet "CP" | Blackout Hr — City Pharmacy |
    | Sheet "CMHL" | Blackout Hr — City Mart Holdings |
    | Sheet "CFC" | Blackout Hr — City Food Chain |
    | Multiple sheets: CMHL, CP, CFC, PG | Daily Fuel Price |

    **Rename the file to anything** — `data_march.xlsx`, `weekly_report.xlsx` — it still works.
    """)

    st.markdown("#### 🔤 Generator Name Typos → Auto-fixed")
    fixes = pd.DataFrame({
        "In Your Excel": ["KHOLER-550", "KOHLER -550", "HIMONISA-200", "550 KVA - G1", "POWER MAX-150"],
        "System Converts To": ["KOHLER-550", "KOHLER-550", "HIMOINSA-200", "550KVA-G1", "POWERMAX-150"],
    })
    st.dataframe(fixes, use_container_width=True, hide_index=True)

    st.markdown("#### 📊 Dynamic Columns → Auto-detected")
    st.markdown("""
    Each day adds 4-5 new columns to your Excel. The parser:
    1. Scans Row 2 for **all date values** (any number of dates)
    2. Reads 4 sub-columns per date: Run Hr, Used, Balance, Blackout
    3. **No code change needed** when new days are added
    """)

    st.markdown("#### 🧹 Messy Values → Cleaned")
    cleaning = pd.DataFrame({
        "In Your Excel": ["-", "X", "#DIV/0!", "empty cell", "2.2666667", "5614 (blackout)"],
        "System Action": ["→ NULL", "→ NULL", "→ NULL", "→ NULL", "→ 2.27 (kept)", "→ REJECTED (>24 hrs)"],
    })
    st.dataframe(cleaning, use_container_width=True, hide_index=True)

    st.markdown("#### 🔄 Re-upload = Update, Not Duplicate")
    st.markdown("""
    - Same date + same generator → **overwritten** with new values
    - New dates → **added**
    - New sites/generators → **auto-created**
    - Upload the same file 10 times → still only 1 copy in database
    """)

# ═══════════════════════════════════════════════════════════════════════════
# TAB: History
# ═══════════════════════════════════════════════════════════════════════════
elif selected == "📋 History":
    st.markdown("### Upload & Database History")
    with get_db() as conn:
        uploads = pd.read_sql_query("""
            SELECT filename, file_type, sector_id, rows_parsed, rows_accepted,
                   rows_rejected, date_range_start, date_range_end, uploaded_at
            FROM upload_history ORDER BY uploaded_at DESC
        """, conn)
        totals = {}
        for t in ["sites", "generators", "daily_operations", "fuel_purchases"]:
            totals[t] = conn.execute(f"SELECT COUNT(*) FROM {t}").fetchone()[0]

    c1, c2, c3, c4 = st.columns(4)
    with c1:
        ui.metric_card(title="Sites", content=str(totals["sites"]), description="in database", key="mc_h_sites")
    with c2:
        ui.metric_card(title="Generators", content=str(totals["generators"]), description="in database", key="mc_h_gens")
    with c3:
        ui.metric_card(title="Daily Records", content=str(totals["daily_operations"]), description="operations", key="mc_h_ops")
    with c4:
        ui.metric_card(title="Fuel Prices", content=str(totals["fuel_purchases"]), description="purchases", key="mc_h_fuel")

    if not uploads.empty:
        st.dataframe(uploads, use_container_width=True, hide_index=True)
    else:
        st.info("No uploads yet.")
