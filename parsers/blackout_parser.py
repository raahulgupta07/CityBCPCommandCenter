"""
Parser for Blackout Hr_ *.xlsx files (CP, CMHL, CFC).

Handles:
- Dynamic column detection from Row 2/3 headers (resilient to new columns)
- Dynamic date sub-column detection (Gen Run Hr, Daily Used, Tank, Blackout)
- Multi-generator per site with sector-prefixed site_ids
- Merged cells, dashes, empty rows
- Sector-specific variations (CMHL has no blackout_hr)
"""
import re
import openpyxl
from datetime import datetime

from parsers.base_parser import clean_numeric, parse_date_from_cell, validate_range
from parsers.name_normalizer import normalize_generator_name, extract_kva_from_model
from config.settings import VALIDATION, SECTORS


# ─── Header keyword matchers (case-insensitive, partial match) ────────────
# Static columns: scan Row 2 + Row 3 headers before the first date column
STATIC_COL_PATTERNS = {
    "seq":              [r"^(sr|no|seq|စဥ်)"],
    "business_sector":  [r"^sector$"],
    "company":          [r"^company$"],
    "store_code":       [r"^store$", r"^site$", r"^center\s*name", r"^center\b(?!.*cost)"],
    "site_name":        [r"cost\s*center\s*name", r"location"],
    "cost_code":        [r"cost\s*center\s*code"],
    "type":             [r"type.*regular", r"regular.*lng"],
    "fuel_type":        [r"fuel\s*type", r"ဆီအမျိုးအစား", r"pd.*hsd"],
    "supplier":         [r"purchased?\s*comp", r"ဝယ်ယူ", r"supplier"],
    "model":            [r"machine\s*power", r"^kva$"],
    "consumption":      [r"consump.*per\s*hour", r"^liter$"],
}

# Date sub-columns: scan Row 3 headers within a date group
DATE_SUB_PATTERNS = {
    "gen_run_hr":          [r"gen(erator)?\s*run\s*hr"],
    "daily_used_liters":   [r"daily\s*used"],
    "spare_tank_balance":  [r"spare\s*tank", r"tank.*balance", r"drum\s*balance"],
    "blackout_hr":         [r"blackout\s*hr"],
}


def _match_header(cell_val, patterns):
    """Check if a cell value matches any of the given regex patterns."""
    if cell_val is None:
        return False
    # Normalize: collapse all whitespace (newlines, tabs, spaces) into single space
    text = re.sub(r'\s+', ' ', str(cell_val)).strip().lower()
    if not text:
        return False
    for pat in patterns:
        if re.search(pat, text, re.IGNORECASE):
            return True
    return False


def _find_static_columns(ws, first_date_col):
    """
    Scan Row 2 and Row 3 (columns 1 to first_date_col-1) for known headers.
    Returns dict of field_name → column_index (1-based).
    """
    cols = {}
    for col_idx in range(1, first_date_col):
        r2 = ws.cell(row=2, column=col_idx).value
        r3 = ws.cell(row=3, column=col_idx).value
        for field, patterns in STATIC_COL_PATTERNS.items():
            if field in cols:
                continue  # already found
            if _match_header(r2, patterns) or _match_header(r3, patterns):
                cols[field] = col_idx
    return cols


def _find_date_sub_offsets(ws, first_date_col, cols_per_date):
    """
    Scan Row 3 within the first date group to find sub-column offsets.
    Returns dict of field_name → offset (0-based from date_col_start).
    """
    offsets = {}
    for offset in range(cols_per_date):
        val = ws.cell(row=3, column=first_date_col + offset).value
        for field, patterns in DATE_SUB_PATTERNS.items():
            if field in offsets:
                continue
            if _match_header(val, patterns):
                offsets[field] = offset
    return offsets


def parse_blackout_file(filepath, sector_id):
    """
    Parse a Blackout Hr_ Excel file and return structured data.

    Returns:
        {
            "generators": [{site_id, site_name, site_type, model_name, model_name_raw,
                            power_kva, consumption_per_hour, fuel_type, supplier}],
            "daily_data": [{site_id, model_name_raw, date, gen_run_hr,
                            daily_used_liters, spare_tank_balance, blackout_hr}],
            "errors": [str],
            "warnings": [str],
            "dates_found": [str],
            "columns_detected": dict,  # for debugging
        }
    """
    wb = openpyxl.load_workbook(str(filepath), data_only=True)

    # Try to find the correct sheet
    if sector_id in wb.sheetnames:
        ws = wb[sector_id]
    else:
        ws = wb[wb.sheetnames[0]]

    # Auto-detect blackout column from Excel rather than relying on config
    # Config is kept as fallback but actual column detection overrides it
    has_blackout = True  # Always try to detect; if column not found, offset won't exist

    result = {
        "generators": [],
        "daily_data": [],
        "errors": [],
        "warnings": [],
        "dates_found": [],
        "columns_detected": {},
    }

    # ─── Step 1: Find date columns in Row 2 ──────────────────────────────
    date_columns = []  # [(date_str, col_index)]
    for col_idx in range(1, ws.max_column + 1):
        cell_val = ws.cell(row=2, column=col_idx).value
        date_str = parse_date_from_cell(cell_val)
        if date_str:
            date_columns.append((date_str, col_idx))

    if not date_columns:
        result["errors"].append("No date columns found in Row 2")
        wb.close()
        return result

    result["dates_found"] = [d[0] for d in date_columns]
    first_date_col = date_columns[0][1]

    # ─── Step 2: Determine sub-column layout per date ────────────────────
    if len(date_columns) >= 2:
        cols_per_date = date_columns[1][1] - date_columns[0][1]
    else:
        # Estimate from row 3 headers
        count = 0
        for c in range(first_date_col, min(first_date_col + 7, ws.max_column + 1)):
            if ws.cell(row=3, column=c).value:
                count += 1
        cols_per_date = max(count + 1, 4)

    # ─── Step 3: Dynamic column detection from headers ───────────────────
    static_cols = _find_static_columns(ws, first_date_col)
    date_sub_offsets = _find_date_sub_offsets(ws, first_date_col, cols_per_date)

    # Store for debugging
    result["columns_detected"] = {
        "static": static_cols,
        "date_sub_offsets": date_sub_offsets,
        "first_date_col": first_date_col,
        "cols_per_date": cols_per_date,
    }

    # Resolve static column indices with fallbacks
    COL_SEQ = static_cols.get("seq", 1)
    COL_SITE = static_cols.get("store_code", static_cols.get("site_name", 2))
    COL_SITE_NAME = static_cols.get("site_name")  # full name (may be None in old format)
    COL_TYPE = static_cols.get("type")
    COL_FUEL = static_cols.get("fuel_type")
    COL_SUPPLIER = static_cols.get("supplier")
    COL_MODEL = static_cols.get("model")
    COL_CONSUMPTION = static_cols.get("consumption")

    # Has separate store code + site name columns?
    has_separate_name = (
        "store_code" in static_cols
        and "site_name" in static_cols
        and static_cols["store_code"] != static_cols["site_name"]
    )

    # Resolve date sub-column offsets with fallbacks
    OFF_RUN_HR = date_sub_offsets.get("gen_run_hr", 0)
    OFF_DAILY_USED = date_sub_offsets.get("daily_used_liters", 1)
    OFF_TANK = date_sub_offsets.get("spare_tank_balance", 2)
    OFF_BLACKOUT = date_sub_offsets.get("blackout_hr", 3)

    # ─── Step 4: Find data start row ─────────────────────────────────────
    data_start_row = None
    for row_idx in range(1, min(20, ws.max_row + 1)):
        cell_val = ws.cell(row=row_idx, column=COL_SEQ).value
        if cell_val is not None:
            try:
                int(float(str(cell_val)))
                data_start_row = row_idx
                break
            except (ValueError, TypeError):
                continue

    if data_start_row is None:
        result["errors"].append("Could not find data start row (no numeric seq# found)")
        wb.close()
        return result

    # ─── Step 5: Parse each data row ─────────────────────────────────────
    seen_generators = set()
    last_site_code = None    # Short code or name (carry forward for merged cells)
    last_site_full_name = None  # Cost Center Name (carry forward)
    last_cost_center_code = None  # Cost Center Code (carry forward)
    last_site_type = "Regular"
    last_business_sector = None  # Business sector from Excel col (carry forward)
    last_company = None          # Company from Excel col (carry forward)
    _merged_tank = {}      # (site_id, date) → tank_balance
    _merged_blackout = {}  # (site_id, date) → blackout_hr

    # Resolve cost code column
    COL_COST_CODE = static_cols.get("cost_code")
    COL_BIZ_SECTOR = static_cols.get("business_sector")
    COL_COMPANY = static_cols.get("company")

    for row_idx in range(data_start_row, ws.max_row + 1):
        seq_val = ws.cell(row=row_idx, column=COL_SEQ).value

        # Skip non-data rows
        if seq_val is None:
            continue
        try:
            int(float(str(seq_val)))
        except (ValueError, TypeError):
            continue

        # Read static columns — handle merged cells by carrying forward
        # Business sector + company (carry forward for merged cells)
        if COL_BIZ_SECTOR:
            bsv = ws.cell(row=row_idx, column=COL_BIZ_SECTOR).value
            if bsv and str(bsv).strip() and str(bsv).strip().lower() not in ("", "none"):
                last_business_sector = str(bsv).strip()
        if COL_COMPANY:
            cpv = ws.cell(row=row_idx, column=COL_COMPANY).value
            if cpv and str(cpv).strip() and str(cpv).strip().lower() not in ("", "none"):
                last_company = str(cpv).strip()

        site_val = ws.cell(row=row_idx, column=COL_SITE).value
        if site_val and str(site_val).strip():
            last_site_code = str(site_val).strip()
            # Read type
            if COL_TYPE:
                type_val = ws.cell(row=row_idx, column=COL_TYPE).value
                last_site_type = str(type_val).strip() if type_val and str(type_val).strip() not in ("", "None") else "Regular"
            # Read full name if available
            if has_separate_name and COL_SITE_NAME:
                ccn = ws.cell(row=row_idx, column=COL_SITE_NAME).value
                if ccn and str(ccn).strip():
                    last_site_full_name = str(ccn).strip()
            # Read cost center code
            if COL_COST_CODE:
                cc_val = ws.cell(row=row_idx, column=COL_COST_CODE).value
                if cc_val is not None:
                    last_cost_center_code = str(int(cc_val)) if isinstance(cc_val, float) else str(cc_val).strip()

        if not last_site_code:
            continue

        # Prefix site_id with sector to avoid cross-sector collisions
        # e.g., CP-CM19, CMHL-CMJS, CFC-SBFTY
        site_id = f"{sector_id}-{last_site_code}"
        site_name = last_site_full_name if has_separate_name and last_site_full_name else last_site_code
        site_type = last_site_type

        # Fuel type
        fuel_type = None
        if COL_FUEL:
            fuel_type_raw = ws.cell(row=row_idx, column=COL_FUEL).value
            fuel_type = str(fuel_type_raw).strip() if fuel_type_raw and str(fuel_type_raw).strip() not in ("", "None") else None

        # Supplier (with normalization)
        supplier = None
        if COL_SUPPLIER:
            supplier_raw = ws.cell(row=row_idx, column=COL_SUPPLIER).value
            supplier = str(supplier_raw).strip() if supplier_raw and str(supplier_raw).strip() not in ("", "None") else None
            if supplier:
                s_upper = supplier.upper().replace(" ", "")
                if s_upper == "DENKO":
                    supplier = "Denko"
                elif s_upper == "MOONSUN":
                    supplier = "Moon Sun"

        # Model / Machine Power
        model_name_raw = "UNKNOWN"
        if COL_MODEL:
            model_raw = ws.cell(row=row_idx, column=COL_MODEL).value
            model_name_raw = str(model_raw).strip() if model_raw else "UNKNOWN"
        model_name = normalize_generator_name(model_name_raw)
        power_kva = extract_kva_from_model(model_name_raw)

        # Consumption per hour
        consumption_per_hour = None
        if COL_CONSUMPTION:
            consumption_per_hour = clean_numeric(ws.cell(row=row_idx, column=COL_CONSUMPTION).value)

        # Store generator info (deduplicate by site + raw model name)
        gen_key = (site_id, model_name_raw)
        if gen_key not in seen_generators:
            seen_generators.add(gen_key)
            result["generators"].append({
                "site_id": site_id,
                "site_name": site_name,
                "site_type": site_type if site_type != "None" else "Regular",
                "cost_center_code": last_cost_center_code,
                "business_sector": last_business_sector,
                "company": last_company,
                "model_name": model_name,
                "model_name_raw": model_name_raw,
                "power_kva": power_kva,
                "consumption_per_hour": consumption_per_hour,
                "fuel_type": fuel_type,
                "supplier": supplier,
            })

        # ─── Read daily data for each date column ────────────────────
        for date_str, date_col_start in date_columns:
            gen_run_hr = clean_numeric(ws.cell(row=row_idx, column=date_col_start + OFF_RUN_HR).value)
            daily_used = clean_numeric(ws.cell(row=row_idx, column=date_col_start + OFF_DAILY_USED).value)
            tank_balance_raw = clean_numeric(ws.cell(row=row_idx, column=date_col_start + OFF_TANK).value)
            blackout_raw = None
            if has_blackout and "blackout_hr" in date_sub_offsets:
                blackout_raw = clean_numeric(ws.cell(row=row_idx, column=date_col_start + OFF_BLACKOUT).value)

            # Tank balance and blackout are SITE-LEVEL values (merged across generators).
            # Store them ONLY on the first generator row that has the value.
            # Other generator rows at the same site get None to avoid double-counting
            # when someone does SUM() across the raw daily_operations table.
            cache_key = (site_id, date_str)

            if tank_balance_raw is not None:
                # First row with the value — record it
                if cache_key not in _merged_tank:
                    _merged_tank[cache_key] = tank_balance_raw
                    tank_balance = tank_balance_raw
                else:
                    # Already stored on a previous generator row — don't duplicate
                    tank_balance = None
            else:
                # Merged cell (None from openpyxl) — check if first row already stored it
                if cache_key not in _merged_tank:
                    tank_balance = None  # genuinely missing
                else:
                    tank_balance = None  # already stored on first gen row

            if blackout_raw is not None:
                if cache_key not in _merged_blackout:
                    _merged_blackout[cache_key] = blackout_raw
                    blackout = blackout_raw
                else:
                    blackout = None
            else:
                if cache_key not in _merged_blackout:
                    blackout = None
                else:
                    blackout = None

            # Skip row-date if no generator-level data AND no site-level data
            if all(v is None for v in [gen_run_hr, daily_used, tank_balance, blackout]):
                # But still keep the row if gen_run_hr or daily_used has data
                # (a generator with 0 run hours is still valid data)
                continue

            # Validate
            warnings_for_row = []

            gen_run_hr, w, rejected = validate_range(gen_run_hr, "gen_run_hr", VALIDATION)
            if rejected:
                result["errors"].append(f"Row {row_idx}, {date_str}: {w}")
                continue
            if w:
                warnings_for_row.append(w)

            daily_used, w, rejected = validate_range(daily_used, "daily_used_liters", VALIDATION)
            if w:
                warnings_for_row.append(w)

            tank_balance, w, rejected = validate_range(tank_balance, "spare_tank_balance", VALIDATION)
            if w:
                warnings_for_row.append(w)

            if blackout is not None:
                blackout, w, rejected = validate_range(blackout, "blackout_hr", VALIDATION)
                if rejected:
                    result["warnings"].append(f"Row {row_idx} ({site_name}), {date_str}: {w} — blackout value ignored, other data kept")
                    blackout = None
                if w and not rejected:
                    warnings_for_row.append(w)

            # Capture raw blackout text notes
            if has_blackout and "blackout_hr" in date_sub_offsets:
                raw_blackout_val = ws.cell(row=row_idx, column=date_col_start + OFF_BLACKOUT).value
                if raw_blackout_val and isinstance(raw_blackout_val, str) and raw_blackout_val.strip():
                    result["warnings"].append(f"Row {row_idx} ({site_name}), {date_str}: Blackout note: '{raw_blackout_val.strip()}'")

            if warnings_for_row:
                result["warnings"].extend(
                    [f"Row {row_idx} ({site_name}/{model_name_raw}), {date_str}: {w}" for w in warnings_for_row]
                )

            result["daily_data"].append({
                "site_id": site_id,
                "model_name_raw": model_name_raw,
                "date": date_str,
                "gen_run_hr": gen_run_hr,
                "daily_used_liters": daily_used,
                "spare_tank_balance": tank_balance,
                "blackout_hr": blackout,
            })

    wb.close()
    return result
