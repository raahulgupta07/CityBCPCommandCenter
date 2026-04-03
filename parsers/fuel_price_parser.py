"""
Parser for Daily Fuel Price.xlsx (4 sheets: CMHL, CP, CFC, PG).

Each sheet layout:
  Row 1: Title
  Row 2: Region markers (YGN cols 3-8, MDY cols 9-14)
  Row 3: Headers (Sector, Date, Supplier, Qty, Price PD, Supplier, Qty, Price HSD, ...)
  Row 4+: Data
"""
import openpyxl

from parsers.base_parser import clean_numeric, parse_date_from_cell, clean_value


def parse_fuel_price_file(filepath):
    """
    Parse Daily Fuel Price.xlsx and return structured data for all sheets.

    Returns:
        {
            "purchases": [
                {
                    "sector_id": str, "date": str, "region": str,
                    "supplier": str, "fuel_type": str,
                    "quantity_liters": float, "price_per_liter": float,
                }
            ],
            "errors": [str],
            "warnings": [str],
        }
    """
    wb = openpyxl.load_workbook(str(filepath), data_only=True)

    result = {
        "purchases": [],
        "errors": [],
        "warnings": [],
    }

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sector_id = sheet_name.strip().upper()
        if sector_id not in ("CP", "CMHL", "CFC", "PG"):
            result["warnings"].append(f"Unknown sheet '{sheet_name}', skipping")
            continue

        # Detect layout: find which column has the date
        # Old format: Sector(1), Date(2), data from col 3
        # New format: Sector(1), Company(2), Date(3), data from col 4
        date_col = None
        data_offset = 0
        for r in range(3, min(10, ws.max_row + 1)):
            # Try col 2 first (old format)
            d2 = parse_date_from_cell(ws.cell(row=r, column=2).value)
            if d2:
                date_col = 2
                data_offset = 0  # data starts at col 3
                break
            # Try col 3 (new format with Company column)
            d3 = parse_date_from_cell(ws.cell(row=r, column=3).value)
            if d3:
                date_col = 3
                data_offset = 1  # data starts at col 4
                break

        if date_col is None:
            date_col = 3  # default to new format
            data_offset = 1

        # Layout (with offset):
        # YGN PD: cols 3+off, 4+off, 5+off (Supplier, Qty, Price)
        # YGN HSD: cols 6+off, 7+off, 8+off
        # MDY PD: cols 9+off, 10+off, 11+off
        # MDY HSD: cols 12+off, 13+off, 14+off

        data_start = None
        for r in range(3, min(10, ws.max_row + 1)):
            date_val = parse_date_from_cell(ws.cell(row=r, column=date_col).value)
            if date_val:
                data_start = r
                break
        if data_start is None:
            data_start = 4

        o = data_offset
        for row_idx in range(data_start, ws.max_row + 1):
            date_str = parse_date_from_cell(ws.cell(row=row_idx, column=date_col).value)
            if not date_str:
                continue

            # YGN PD
            _add_purchase(result, sector_id, date_str, "YGN", "PD",
                          ws.cell(row=row_idx, column=3+o).value,
                          ws.cell(row=row_idx, column=4+o).value,
                          ws.cell(row=row_idx, column=5+o).value)
            # YGN HSD
            _add_purchase(result, sector_id, date_str, "YGN", "HSD",
                          ws.cell(row=row_idx, column=6+o).value,
                          ws.cell(row=row_idx, column=7+o).value,
                          ws.cell(row=row_idx, column=8+o).value)
            # MDY PD
            _add_purchase(result, sector_id, date_str, "MDY", "PD",
                          ws.cell(row=row_idx, column=9+o).value,
                          ws.cell(row=row_idx, column=10+o).value,
                          ws.cell(row=row_idx, column=11+o).value)
            # MDY HSD
            _add_purchase(result, sector_id, date_str, "MDY", "HSD",
                          ws.cell(row=row_idx, column=12+o).value,
                          ws.cell(row=row_idx, column=13+o).value,
                          ws.cell(row=row_idx, column=14+o).value)

    wb.close()
    return result


def _add_purchase(result, sector_id, date_str, region, fuel_type,
                  supplier_val, qty_val, price_val):
    """Add a fuel purchase record if there's meaningful data."""
    supplier = clean_value(supplier_val)
    quantity = clean_numeric(qty_val)
    price = clean_numeric(price_val)

    # Skip if no supplier AND no price (completely empty entry)
    if supplier is None and price is None:
        return

    # Clean and normalize supplier name
    if isinstance(supplier, (int, float)):
        supplier = None  # numeric in supplier column = data error
    if supplier:
        supplier = str(supplier).strip()
        # Normalize known supplier name variants
        s_upper = supplier.upper().replace(" ", "")
        if s_upper == "DENKO":
            supplier = "Denko"
        elif s_upper == "MOONSUN":
            supplier = "Moon Sun"

    result["purchases"].append({
        "sector_id": sector_id,
        "date": date_str,
        "region": region,
        "supplier": supplier,
        "fuel_type": fuel_type,
        "quantity_liters": quantity,
        "price_per_liter": price,
    })
