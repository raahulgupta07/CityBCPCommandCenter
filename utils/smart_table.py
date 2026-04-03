"""
HTML Smart Tables with dark header, severity badges, conditional coloring,
progress bars, and Excel download button.
"""
import io
import pandas as pd
import streamlit as st


def render_smart_table(df, title=None, severity_col=None, highlight_cols=None,
                       max_height=500, bar_col=None, download=True):
    """
    Render a styled HTML table with dark header and optional Excel download.

    Args:
        df: DataFrame to render
        title: Optional table title (shown in dark header)
        severity_col: Column name containing severity values (CRITICAL/WARNING/etc.)
        highlight_cols: Dict of {col_name: {"good": "high"|"low", "thresholds": [warn, crit]}}
        max_height: Max table height in px
        bar_col: Column name to show as progress bar (0-100)
        download: Show Excel download button (default True)
    """
    if df is None or df.empty:
        st.info("No data available")
        return

    rows_html = []
    for _, row in df.iterrows():
        cells = []
        for col in df.columns:
            val = row[col]
            style = ""
            content = _format_value(val)

            # Severity badge
            if col == severity_col and val:
                content = _severity_badge(str(val).upper())

            # Threshold coloring
            elif highlight_cols and col in highlight_cols:
                style = _threshold_style(val, highlight_cols[col])

            # Progress bar
            elif col == bar_col and val is not None:
                content = _progress_bar(val)

            cells.append(f'<td style="{style}">{content}</td>')

        rows_html.append(f'<tr>{"".join(cells)}</tr>')

    headers = "".join(f"<th>{col}</th>" for col in df.columns)
    title_text = f"{title} ({len(df)} rows)" if title else f"{len(df)} rows"

    uid = f"tbl_{id(df)}_{hash(title or '')}"

    html = f"""
    <style>
    .stbl-wrap-{uid} {{
        max-height: {max_height}px;
        overflow: auto;
        border-radius: 0 0 12px 12px;
        border: 1px solid #e5e7eb;
        border-top: none;
    }}
    .stbl-{uid} {{
        width: 100%;
        border-collapse: collapse;
        font-size: 13px;
    }}
    .stbl-{uid} th {{
        position: sticky;
        top: 0;
        background: #1e293b;
        padding: 10px 12px;
        text-align: left;
        font-weight: 600;
        color: #cbd5e1;
        white-space: nowrap;
        font-size: 12px;
    }}
    .stbl-{uid} td {{
        padding: 7px 12px;
        border-bottom: 1px solid #f3f4f6;
        color: #374151;
        font-size: 13px;
    }}
    .stbl-{uid} tr:nth-child(even) {{
        background: #fafbfc;
    }}
    .stbl-{uid} tr:hover {{
        background: #eff6ff;
    }}
    .severity-badge {{
        display: inline-block;
        padding: 2px 8px;
        border-radius: 12px;
        font-size: 11px;
        font-weight: 600;
        text-transform: uppercase;
    }}
    </style>
    <div style="border:1px solid #e5e7eb;border-radius:12px;margin:8px 0;overflow:hidden">
        <div style="background:#0f172a;color:white;padding:10px 16px;font-weight:700;font-size:14px">
            {title_text}
        </div>
        <div class="stbl-wrap-{uid}">
            <table class="stbl-{uid}">
                <thead><tr>{headers}</tr></thead>
                <tbody>{"".join(rows_html)}</tbody>
            </table>
        </div>
    </div>
    """
    st.html(html)

    # Excel download button
    if download:
        _render_download(df, title or "table")


def _render_download(df, title, color_rules=None):
    """
    Render an Excel download button with colored cells.

    Args:
        df: DataFrame to export
        title: Filename and button label
        color_rules: Optional dict of {col_name: metric_key} for heatmap coloring.
                     metric_key is one of: diesel_price, blackout_hr, expense_pct, buffer_days
    """
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from config.settings import HEATMAP_THRESHOLDS, HEATMAP_COLORS

    buf = io.BytesIO()
    clean_name = "".join(c if c.isalnum() or c in " _-" else "_" for c in title).strip()[:50]

    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Data")
        ws = writer.sheets["Data"]

        # Style header row — dark background, white text
        header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        thin_border = Border(
            bottom=Side(style="thin", color="E5E7EB"),
            right=Side(style="thin", color="E5E7EB"),
        )

        for col_idx in range(1, len(df.columns) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # No background fills — icons convey status. Keep white background.
        _no_fill = PatternFill()
        _fill_map = {
            "green":  _no_fill,
            "yellow": _no_fill,
            "amber":  _no_fill,
            "red":    _no_fill,
            "gray":   _no_fill,
        }
        _font_map = {
            "green":  Font(bold=True, color="166534"),
            "yellow": Font(bold=True, color="854D0E"),
            "amber":  Font(bold=True, color="9A3412"),
            "red":    Font(bold=True, color="991B1B"),
            "gray":   Font(color="6B7280"),
        }

        def _get_color_name(value, metric_key):
            if value is None or (isinstance(value, float) and value != value):
                return "gray"
            cfg = HEATMAP_THRESHOLDS[metric_key]
            if cfg.get("reverse"):
                if value >= cfg["green_min"]: return "green"
                if value >= cfg["yellow_min"]: return "yellow"
                if value >= cfg["amber_min"]: return "amber"
                return "red"
            else:
                if value < cfg["green_max"]: return "green"
                if value < cfg["yellow_max"]: return "yellow"
                if value <= cfg["amber_max"]: return "amber"
                return "red"

        # Apply heatmap colors to specified columns
        if color_rules:
            col_names = list(df.columns)
            for col_name, metric_key in color_rules.items():
                if col_name not in col_names:
                    continue
                col_idx = col_names.index(col_name) + 1  # 1-based
                for row_idx in range(2, len(df) + 2):  # skip header
                    cell = ws.cell(row=row_idx, column=col_idx)
                    try:
                        val = float(cell.value) if cell.value is not None else None
                    except (ValueError, TypeError):
                        val = None
                    cname = _get_color_name(val, metric_key)
                    cell.fill = _fill_map[cname]
                    cell.font = _font_map[cname]
                    cell.alignment = Alignment(horizontal="center")
                    # Add icon to cell value for readability without colors
                    _icons = {"green": "🟢", "yellow": "🟡", "amber": "🟠", "red": "🔴", "gray": "⚪"}
                    icon = _icons.get(cname, "")
                    if val is not None:
                        if metric_key == "diesel_price":
                            cell.value = f"{icon} {val:,.0f}"
                        elif metric_key == "blackout_hr":
                            cell.value = f"{icon} {val:.1f}"
                        elif metric_key == "expense_pct":
                            cell.value = f"{icon} {val:.2f}%"
                        elif metric_key == "buffer_days":
                            cell.value = f"{icon} {val:.1f}"
                        else:
                            cell.value = f"{icon} {val}"
                    else:
                        cell.value = f"{icon} N/A"

        # Apply severity badge colors (for render_smart_table tables)
        if not color_rules:
            sev_fills = {
                "CRITICAL": PatternFill(start_color="FECACA", end_color="FECACA", fill_type="solid"),
                "WARNING":  PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"),
                "HEALTHY":  PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid"),
                "SAFE":     PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid"),
                "FULL":     PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid"),
                "REDUCED":  PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"),
                "CLOSE":    PatternFill(start_color="FECACA", end_color="FECACA", fill_type="solid"),
                "IMMEDIATE": PatternFill(start_color="FECACA", end_color="FECACA", fill_type="solid"),
                "TODAY":    PatternFill(start_color="FECACA", end_color="FECACA", fill_type="solid"),
                "TOMORROW": PatternFill(start_color="FFEDD5", end_color="FFEDD5", fill_type="solid"),
                "A": PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid"),
                "B": PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid"),
                "C": PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"),
                "D": PatternFill(start_color="FFEDD5", end_color="FFEDD5", fill_type="solid"),
                "F": PatternFill(start_color="FECACA", end_color="FECACA", fill_type="solid"),
            }
            for row_idx in range(2, len(df) + 2):
                for col_idx in range(1, len(df.columns) + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    val = str(cell.value).strip().upper() if cell.value else ""
                    if val in sev_fills:
                        cell.fill = sev_fills[val]
                        cell.font = Font(bold=True)

        # Auto-fit column widths
        for col_idx in range(1, len(df.columns) + 1):
            max_len = len(str(ws.cell(row=1, column=col_idx).value or ""))
            for row_idx in range(2, min(len(df) + 2, 52)):
                val_len = len(str(ws.cell(row=row_idx, column=col_idx).value or ""))
                if val_len > max_len:
                    max_len = val_len
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 3, 30)

        # Alternate row coloring
        even_fill = PatternFill(start_color="FAFBFC", end_color="FAFBFC", fill_type="solid")
        for row_idx in range(2, len(df) + 2):
            if row_idx % 2 == 0:
                for col_idx in range(1, len(df.columns) + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if cell.fill.start_color.index == "00000000":  # only if no color already
                        cell.fill = even_fill

    buf.seek(0)
    st.download_button(
        label=f"Download {clean_name}.xlsx",
        data=buf,
        file_name=f"{clean_name}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key=f"dl_{id(df)}_{hash(title)}",
    )


# ─── Helpers ─────────────────────────────────────────────────────────────────

SEVERITY_STYLES = {
    "CRITICAL": {"bg": "#fef2f2", "color": "#dc2626"},
    "HIGH": {"bg": "#fff7ed", "color": "#ea580c"},
    "WARNING": {"bg": "#fffbeb", "color": "#d97706"},
    "MEDIUM": {"bg": "#fffbeb", "color": "#d97706"},
    "LOW": {"bg": "#f0fdf4", "color": "#16a34a"},
    "SAFE": {"bg": "#f0fdf4", "color": "#16a34a"},
    "HEALTHY": {"bg": "#f0fdf4", "color": "#16a34a"},
    "NORMAL": {"bg": "#f0fdf4", "color": "#16a34a"},
    "INFO": {"bg": "#eff6ff", "color": "#2563eb"},
    "FULL": {"bg": "#f0fdf4", "color": "#16a34a"},
    "REDUCED": {"bg": "#fffbeb", "color": "#d97706"},
    "GENERATOR_ONLY": {"bg": "#fff7ed", "color": "#ea580c"},
    "CLOSE": {"bg": "#fef2f2", "color": "#dc2626"},
    "NO DATA": {"bg": "#f3f4f6", "color": "#6b7280"},
    "IMMEDIATE": {"bg": "#fef2f2", "color": "#dc2626"},
    "TODAY": {"bg": "#fef2f2", "color": "#dc2626"},
    "TOMORROW": {"bg": "#fff7ed", "color": "#ea580c"},
    "THIS_WEEK": {"bg": "#fffbeb", "color": "#d97706"},
    "A": {"bg": "#f0fdf4", "color": "#16a34a"},
    "B": {"bg": "#eff6ff", "color": "#2563eb"},
    "C": {"bg": "#fffbeb", "color": "#d97706"},
    "D": {"bg": "#fff7ed", "color": "#ea580c"},
    "F": {"bg": "#fef2f2", "color": "#dc2626"},
}


def _severity_badge(value):
    style = SEVERITY_STYLES.get(value, {"bg": "#f3f4f6", "color": "#6b7280"})
    return (
        f'<span class="severity-badge" '
        f'style="background:{style["bg"]};color:{style["color"]}">'
        f'{value}</span>'
    )


def _threshold_style(value, config):
    if value is None or not isinstance(value, (int, float)):
        return ""
    good = config.get("good", "high")
    thresholds = config.get("thresholds", [])
    if len(thresholds) < 2:
        return ""

    warn, crit = thresholds
    if good == "high":
        if value < crit:
            return "background:#fef2f2;color:#dc2626;font-weight:600;"
        elif value < warn:
            return "background:#fffbeb;color:#d97706;"
        else:
            return "color:#16a34a;"
    else:  # good == "low"
        if value > crit:
            return "background:#fef2f2;color:#dc2626;font-weight:600;"
        elif value > warn:
            return "background:#fffbeb;color:#d97706;"
        else:
            return "color:#16a34a;"


def _progress_bar(value):
    try:
        pct = float(value)
    except (TypeError, ValueError):
        return str(value)

    pct = max(0, min(100, pct))
    if pct >= 80:
        color = "#16a34a"
    elif pct >= 50:
        color = "#d97706"
    else:
        color = "#dc2626"

    return (
        f'<div style="display:flex;align-items:center;gap:6px">'
        f'<div style="height:8px;border-radius:4px;flex:1;background:#e5e7eb">'
        f'<div style="height:100%;border-radius:4px;width:{pct}%;background:{color}"></div></div>'
        f'<span style="font-size:11px;min-width:35px">{pct:.0f}%</span>'
        f'</div>'
    )


def _format_value(val):
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return '<span style="color:#9ca3af">—</span>'
    if isinstance(val, float):
        if val >= 1_000_000:
            return f"{val / 1_000_000:.1f}M"
        if val >= 10_000:
            return f"{val:,.0f}"
        if val == int(val):
            return f"{int(val):,}"
        return f"{val:,.1f}"
    if isinstance(val, int):
        return f"{val:,}"
    return str(val)
